import os
import io
import json
from datetime import datetime
from werkzeug.utils import secure_filename
from werkzeug.security import generate_password_hash, check_password_hash
from flask import Flask, render_template, request, redirect, url_for, session, flash, send_file, jsonify
from database import db, Utilisateur, Client, Facture, Operation, Parametres, LigneFacture, Service, Paiement, Evenement, Technicien, EvenementTechnicien

app = Flask(__name__)
app.secret_key = 'ks_production_2026'

UPLOAD_FOLDER = os.path.join('static', 'uploads')
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

app.config['SQLALCHEMY_DATABASE_URI'] = 'sqlite:///ks_production.db'
app.config['SQLALCHEMY_TRACK_MODIFICATIONS'] = False

db.init_app(app)

def allowed_file(filename):
    ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif', 'svg', 'webp'}
    return '.' in filename and filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

# ================================================================
# INITIALISATION BASE DE DONNÉES
# ================================================================
def initialiser_base():
    db.create_all()

    if Utilisateur.query.count() == 0:
        utilisateurs = [
            Utilisateur(username='admin',    password=generate_password_hash('admin2025'),  role='Administrateur', nom_complet='Admin KS Production'),
            Utilisateur(username='caissier', password=generate_password_hash('caisse2025'), role='Caissier',        nom_complet='Caissier KS Production'),
            Utilisateur(username='lecture',  password=generate_password_hash('view2025'),   role='Lecture seule',   nom_complet='Lecture KS Production'),
        ]
        for u in utilisateurs:
            db.session.add(u)
        db.session.commit()

    if Client.query.count() == 0:
        clients = [
            Client(numero='CLI-001', nom='AWUTE Kossi',      adresse='Lomé',      telephone='92117715'),
            Client(numero='CLI-002', nom='KONDO KOFFI',      adresse='Lomé',      telephone='90354873'),
            Client(numero='CLI-003', nom='La Voie au Togo',  adresse='Lomé-Togo', telephone='90018327'),
        ]
        for c in clients:
            db.session.add(c)
        db.session.commit()

    if Operation.query.count() == 0:
        operations = [
            Operation(numero='OPE-001', nom_client='KONDO KOFFI',   service='Sonorisation Mariage',  montant_ttc=236000, type_operation='Recettes', section='Sonorisation', cree_par='admin'),
            Operation(numero='OPE-002', nom_client='La Voie au Togo',service='Location Materiel',     montant_ttc=59000,  type_operation='Recettes', section='Sonorisation', cree_par='admin'),
            Operation(numero='OPE-003', nom_client='La Voie au Togo',service='Sonorisation Concert',  montant_ttc=354000, type_operation='Recettes', section='Sonorisation', cree_par='admin'),
            Operation(numero='OPE-004', nom_client='AWUTE Kossi',    service='Enregistrement Studio', montant_ttc=41300,  type_operation='Recettes', section='Studio',       cree_par='admin'),
            Operation(numero='OPE-005', nom_client='AWUTE PACOME',   service='Salaires / Prestation', montant_ttc=50000,  type_operation='Depenses', section='Studio',       cree_par='admin'),
        ]
        for o in operations:
            db.session.add(o)
        db.session.commit()

    # ✅ CORRECTION : Parametres et Service dans des blocs séparés avec indentation correcte
    if Parametres.query.count() == 0:
        db.session.add(Parametres())
        db.session.commit()

    if Service.query.count() == 0:
        services_defaut = [
            Service(section='Studio',       libelle='Enregistrement Studio',  prix=35000),
            Service(section='Studio',       libelle='Mixage',                 prix=25000),
            Service(section='Studio',       libelle='Mastering',              prix=30000),
            Service(section='Studio',       libelle='Doublage / Voix off',    prix=20000),
            Service(section='Sonorisation', libelle='Sonorisation Événement', prix=150000),
            Service(section='Sonorisation', libelle='Sonorisation Mariage',   prix=200000),
            Service(section='Sonorisation', libelle='Sonorisation Concert',   prix=300000),
            Service(section='Sonorisation', libelle='Location Matériel',      prix=50000),
        ]
        for s in services_defaut:
            db.session.add(s)
        db.session.commit()

    # ── FACTURES DE TEST ──────────────────────────────────────────
    if Facture.query.filter(Facture.type_operation != 'Proforma').count() == 0:
        from datetime import timedelta
        auj = datetime.now()
        factures_test = [
            Facture(numero='FKSP-01032026-001', nom_client='AWUTE Kossi',
                service='Enregistrement Studio', montant_ttc=41300,
                mode_paiement='Espece', etat_paiement='Payer',
                type_operation='Recettes', section='Studio',
                cree_par='admin', montant_paye=41300, reste_du=0,
                date=auj - timedelta(days=40)),
            Facture(numero='FKSP-05032026-002', nom_client='KONDO KOFFI',
                service='Sonorisation Mariage', montant_ttc=236000,
                mode_paiement='Virement', etat_paiement='Payer',
                type_operation='Recettes', section='Sonorisation',
                cree_par='admin', montant_paye=236000, reste_du=0,
                date=auj - timedelta(days=35)),
            Facture(numero='FKSP-10032026-003', nom_client='La Voie au Togo',
                service='Sonorisation Concert', montant_ttc=354000,
                mode_paiement='Cheque', etat_paiement='Partiel',
                type_operation='Recettes', section='Sonorisation',
                cree_par='admin', montant_paye=150000, reste_du=204000,
                date=auj - timedelta(days=20)),
            Facture(numero='FKSP-08032026-004', nom_client='La Voie au Togo',
                service='Location Matériel', montant_ttc=59000,
                mode_paiement='Espece', etat_paiement='Non Payer',
                type_operation='Recettes', section='Sonorisation',
                cree_par='admin', montant_paye=0, reste_du=59000,
                date=auj - timedelta(days=32)),
            Facture(numero='FKSP-22032026-005', nom_client='KONDO KOFFI',
                service='Mixage', montant_ttc=29500,
                mode_paiement='Espece', etat_paiement='Non Payer',
                type_operation='Recettes', section='Studio',
                cree_par='caissier', montant_paye=0, reste_du=29500,
                date=auj - timedelta(days=18)),
            Facture(numero='FKSP-01042026-006', nom_client='AWUTE Kossi',
                service='Mastering', montant_ttc=35400,
                mode_paiement='Mobile Money', etat_paiement='Payer',
                type_operation='Recettes', section='Studio',
                cree_par='admin', montant_paye=35400, reste_du=0,
                date=auj - timedelta(days=9)),
        ]
        for f in factures_test:
            db.session.add(f)
        db.session.commit()
        f_partielle = Facture.query.filter_by(numero='FKSP-10032026-003').first()
        if f_partielle:
            db.session.add(LigneFacture(facture_id=f_partielle.id,
                service='Sonorisation Concert', prix_unitaire=300000,
                quantite=1, montant_ht=300000, montant_ttc=354000))
            db.session.add(Paiement(numero='PAY-001', facture_id=f_partielle.id,
                n_facture=f_partielle.numero, nom_client='La Voie au Togo',
                montant_facture=354000, montant_paye=150000, reste_du=204000,
                mode_paiement='Cheque', etat_facture='Partiel',
                notes='Avance versée', cree_par='admin'))
            db.session.commit()

    # ── PROFORMAS DE TEST ─────────────────────────────────────────
    if Facture.query.filter_by(type_operation='Proforma').count() == 0:
        from datetime import timedelta
        auj = datetime.now()
        for p in [
            Facture(numero='PROF-01042026-001', nom_client='KONDO KOFFI',
                service='Sonorisation Événement', montant_ttc=177000,
                mode_paiement='Espece', etat_paiement='Non Payer',
                type_operation='Proforma', section='Sonorisation',
                cree_par='admin', montant_paye=0, reste_du=177000,
                date=auj - timedelta(days=5)),
            Facture(numero='PROF-05042026-002', nom_client='AWUTE Kossi',
                service='Enregistrement Studio + Mixage', montant_ttc=70800,
                mode_paiement='Virement', etat_paiement='Non Payer',
                type_operation='Proforma', section='Studio',
                cree_par='admin', montant_paye=0, reste_du=70800,
                date=auj - timedelta(days=2)),
        ]:
            db.session.add(p)
        db.session.commit()

    # ── ÉVÉNEMENTS AGENDA DE TEST ─────────────────────────────────
    if Evenement.query.count() == 0:
        from datetime import date as date_type
        f1 = Facture.query.filter_by(numero='FKSP-05032026-002').first()
        for e in [
            Evenement(titre='Sonorisation Mariage KONDO',
                date=date_type(2026, 4, 22), heure_debut='09:00', heure_fin='20:00',
                nom_client='KONDO KOFFI', service='Sonorisation Mariage',
                section='Sonorisation', lieu='Salle des fêtes, Lomé',
                notes='Prévoir 2 enceintes + micro HF', statut='Confirmé',
                cree_par='admin', facture_id=f1.id if f1 else None),
            Evenement(titre='Enregistrement Studio - AWUTE',
                date=date_type(2026, 4, 25), heure_debut='10:00', heure_fin='14:00',
                nom_client='AWUTE Kossi', service='Enregistrement Studio',
                section='Studio', lieu='KS Production Studio',
                notes='Session de 4h - album solo', statut='Confirmé', cree_par='admin'),
            Evenement(titre='Concert La Voie au Togo',
                date=date_type(2026, 4, 29), heure_debut='18:00', heure_fin='23:00',
                nom_client='La Voie au Togo', service='Sonorisation Concert',
                section='Sonorisation', lieu='Palais des Congrès, Lomé',
                notes='Grand concert - vérifier matériel la veille',
                statut='Tentative', cree_par='admin'),
            Evenement(titre='Mixage album - KONDO KOFFI',
                date=date_type(2026, 4, 30), heure_debut='14:00', heure_fin='18:00',
                nom_client='KONDO KOFFI', service='Mixage',
                section='Studio', lieu='KS Production Studio',
                notes='Suite session du 25 avril', statut='Confirmé', cree_par='caissier'),
        ]:
            db.session.add(e)
        db.session.commit()

    # ── TECHNICIENS DE TEST ───────────────────────────────────────
    if Technicien.query.count() == 0:
        for t in [
            Technicien(nom='MENSAH Kodjo',   telephone='91234567', specialite='Sonorisation', statut='Disponible'),
            Technicien(nom='AGBE Komlan',    telephone='93456789', specialite='Studio',       statut='Disponible'),
            Technicien(nom='DOSSOU Afi',     telephone='95678901', specialite='Régie',        statut='Disponible'),
            Technicien(nom='KPODO Edem',     telephone='97890123', specialite='Sonorisation', statut='Inactif'),
        ]:
            db.session.add(t)
        db.session.commit()
        evt1  = Evenement.query.filter_by(titre='Sonorisation Mariage KONDO').first()
        tech1 = Technicien.query.filter_by(nom='MENSAH Kodjo').first()
        tech3 = Technicien.query.filter_by(nom='DOSSOU Afi').first()
        if evt1 and tech1:
            db.session.add(EvenementTechnicien(evenement_id=evt1.id, technicien_id=tech1.id, role='Chef de son'))
        if evt1 and tech3:
            db.session.add(EvenementTechnicien(evenement_id=evt1.id, technicien_id=tech3.id, role='Assistant'))
        db.session.commit()

# ================================================================
# AUTH
# ================================================================

def generer_numero_facture(prefix='FKSP'):
    """Génère un numéro unique FKSP-DDMMYYYY-NNN, séquence réinitialisée chaque année."""
    now    = datetime.now()
    date_str = now.strftime('%d%m%Y')   # ex: 09042026
    annee    = now.strftime('%Y')        # ex: 2026
    # Compter toutes les factures de l'année courante (pas juste du jour)
    nb = Facture.query.filter(
        Facture.numero.like(f'{prefix}-___{annee}-%')
    ).count() + 1
    numero = f"{prefix}-{date_str}-{nb:03d}"
    while Facture.query.filter_by(numero=numero).first():
        nb += 1
        numero = f"{prefix}-{date_str}-{nb:03d}"
    return numero

@app.route('/')
def accueil():
    return redirect(url_for('login'))

@app.route('/login', methods=['GET', 'POST'])
def login():
    erreur = None
    if request.method == 'POST':
        username = request.form['username']
        password = request.form['password']
        user = Utilisateur.query.filter_by(username=username).first()
        if user and check_password_hash(user.password, password):
            session['username'] = user.username
            session['role']     = user.role
            return redirect(url_for('dashboard'))
        else:
            erreur = "Nom d'utilisateur ou mot de passe incorrect."
    params = Parametres.query.first()
    return render_template('login.html', erreur=erreur, params=params)

@app.route('/logout')
def logout():
    session.clear()
    return redirect(url_for('login'))

# ================================================================
# ROUTE DASHBOARD AVEC ALERTES — Remplace l'ancienne route /dashboard
# ================================================================

@app.route('/dashboard')
def dashboard():
    if 'username' not in session:
        return redirect(url_for('login'))

    from sqlalchemy import extract, func
    from datetime import datetime, timedelta
    import json

    # ── STATS GÉNÉRALES ──
    total_recettes = db.session.query(func.sum(Operation.montant_ttc))\
        .filter_by(type_operation='Recettes').scalar() or 0
    total_depenses = db.session.query(func.sum(Operation.montant_ttc))\
        .filter_by(type_operation='Depenses').scalar() or 0
    solde_net      = total_recettes - total_depenses
    total_impayes  = db.session.query(func.sum(Facture.reste_du))\
        .filter(Facture.type_operation != 'Proforma').scalar() or 0

    nb_factures   = Facture.query.filter(Facture.type_operation != 'Proforma').count()
    nb_payees     = Facture.query.filter_by(etat_paiement='Payer').filter(Facture.type_operation != 'Proforma').count()
    nb_impayes    = Facture.query.filter_by(etat_paiement='Non Payer').filter(Facture.type_operation != 'Proforma').count()
    nb_partielles = Facture.query.filter_by(etat_paiement='Partiel').filter(Facture.type_operation != 'Proforma').count()
    taux_recouvrement = round((nb_payees / nb_factures * 100) if nb_factures > 0 else 0, 1)

    # ── ALERTES IMPAYÉS ──
    today = datetime.now()

    # Factures impayées depuis plus de 30 jours
    seuil_30j  = today - timedelta(days=30)
    seuil_15j  = today - timedelta(days=15)
    seuil_7j   = today - timedelta(days=7)

    factures_impayees_30j = Facture.query.filter(
        Facture.etat_paiement.in_(['Non Payer', 'Partiel']),
        Facture.type_operation != 'Proforma',
        Facture.date <= seuil_30j
    ).order_by(Facture.date.asc()).all()

    factures_impayees_15j = Facture.query.filter(
        Facture.etat_paiement.in_(['Non Payer', 'Partiel']),
        Facture.type_operation != 'Proforma',
        Facture.date <= seuil_15j,
        Facture.date > seuil_30j
    ).order_by(Facture.date.asc()).all()

    factures_impayees_7j = Facture.query.filter(
        Facture.etat_paiement.in_(['Non Payer', 'Partiel']),
        Facture.type_operation != 'Proforma',
        Facture.date <= seuil_7j,
        Facture.date > seuil_15j
    ).order_by(Facture.date.asc()).all()

    # Montant total en retard
    montant_retard_30j = sum(f.reste_du for f in factures_impayees_30j)
    montant_retard_15j = sum(f.reste_du for f in factures_impayees_15j)

    # Construire les alertes
    alertes = []

    # Alerte solde négatif
    if solde_net < 0:
        alertes.append({
            'type'   : 'danger',
            'icone'  : 'bi-exclamation-triangle-fill',
            'titre'  : '⚠️ Solde négatif !',
            'message': f'Votre solde est négatif : {solde_net:,.0f} FCFA. Les dépenses dépassent les recettes.'.replace(',', ' '),
            'lien'   : '/rapport',
            'lien_txt': 'Voir le rapport',
        })

    # Alerte factures en retard > 30 jours
    if factures_impayees_30j:
        alertes.append({
            'type'   : 'danger',
            'icone'  : 'bi-clock-history',
            'titre'  : f'{len(factures_impayees_30j)} facture(s) impayée(s) depuis plus de 30 jours',
            'message': f'Montant total en retard : {montant_retard_30j:,.0f} FCFA'.replace(',', ' '),
            'lien'   : '/factures?etat=Non+Payer',
            'lien_txt': 'Voir les factures',
            'factures': factures_impayees_30j[:3],  # Max 3 exemples
        })

    # Alerte factures en retard 15-30 jours
    if factures_impayees_15j:
        alertes.append({
            'type'   : 'warning',
            'icone'  : 'bi-exclamation-circle-fill',
            'titre'  : f'{len(factures_impayees_15j)} facture(s) impayée(s) depuis 15 à 30 jours',
            'message': f'Montant : {montant_retard_15j:,.0f} FCFA — Pensez à relancer ces clients.'.replace(',', ' '),
            'lien'   : '/factures?etat=Non+Payer',
            'lien_txt': 'Voir les factures',
        })

    # Alerte factures en retard 7-15 jours
    if factures_impayees_7j:
        alertes.append({
            'type'   : 'info',
            'icone'  : 'bi-info-circle-fill',
            'titre'  : f'{len(factures_impayees_7j)} facture(s) impayée(s) depuis 7 à 15 jours',
            'message': 'Ces factures arrivent bientôt à échéance.',
            'lien'   : '/factures?etat=Non+Payer',
            'lien_txt': 'Voir les factures',
        })

    # ── RÉPARTITION PAR SECTION ──
    sections_data = db.session.query(
        Operation.section,
        func.sum(Operation.montant_ttc)
    ).filter_by(type_operation='Recettes')\
     .group_by(Operation.section).all()

    sections_labels = [s[0] or 'Autre' for s in sections_data]
    sections_values = [float(s[1]) for s in sections_data]

    # ── ÉVOLUTION MENSUELLE (6 derniers mois) ──
    mois_labels, mois_recettes, mois_depenses = [], [], []
    for i in range(5, -1, -1):
        d = today - timedelta(days=i * 30)
        mois_labels.append(d.strftime('%b %y'))
        rec = db.session.query(func.sum(Operation.montant_ttc)).filter(
            Operation.type_operation == 'Recettes',
            extract('month', Operation.date) == d.month,
            extract('year',  Operation.date) == d.year
        ).scalar() or 0
        dep = db.session.query(func.sum(Operation.montant_ttc)).filter(
            Operation.type_operation == 'Depenses',
            extract('month', Operation.date) == d.month,
            extract('year',  Operation.date) == d.year
        ).scalar() or 0
        mois_recettes.append(float(rec))
        mois_depenses.append(float(dep))

    # ── TOP 5 CLIENTS ──
    top_clients = db.session.query(
        Operation.nom_client,
        func.sum(Operation.montant_ttc).label('total')
    ).filter_by(type_operation='Recettes')\
     .group_by(Operation.nom_client)\
     .order_by(func.sum(Operation.montant_ttc).desc())\
     .limit(5).all()

    top_clients_labels = [c[0] or 'Inconnu' for c in top_clients]
    top_clients_values = [float(c[1]) for c in top_clients]

    # ── RÉPARTITION FACTURES ──
    factures_labels = ['Payées', 'Non Payées', 'Partielles']
    factures_values = [
        Facture.query.filter_by(etat_paiement='Payer').filter(Facture.type_operation != 'Proforma').count(),
        Facture.query.filter_by(etat_paiement='Non Payer').filter(Facture.type_operation != 'Proforma').count(),
        Facture.query.filter_by(etat_paiement='Partiel').filter(Facture.type_operation != 'Proforma').count(),
    ]

    # ── 8 DERNIÈRES OPÉRATIONS ──
    dernieres_ops = Operation.query.order_by(Operation.date.desc()).limit(8).all()

    return render_template('dashboard.html',
        username      = session['username'],
        role          = session['role'],
        date_maj      = datetime.now().strftime('%d/%m/%Y %H:%M'),

        # Stats
        total_recettes    = f"{total_recettes:,.0f}".replace(',', ' '),
        total_depenses    = f"{total_depenses:,.0f}".replace(',', ' '),
        solde_net         = f"{solde_net:,.0f}".replace(',', ' '),
        solde_positif     = solde_net >= 0,
        total_impayes     = f"{total_impayes:,.0f}".replace(',', ' '),
        nb_factures       = nb_factures,
        nb_payees         = nb_payees,
        nb_impayes        = nb_impayes,
        nb_partielles     = nb_partielles,
        taux_recouvrement = taux_recouvrement,

        # Alertes
        alertes = alertes,

        # Graphiques
        sections_labels    = json.dumps(sections_labels),
        sections_values    = json.dumps(sections_values),
        mois_labels        = json.dumps(mois_labels),
        mois_recettes      = json.dumps(mois_recettes),
        mois_depenses      = json.dumps(mois_depenses),
        top_clients_labels = json.dumps(top_clients_labels),
        top_clients_values = json.dumps(top_clients_values),
        factures_labels    = json.dumps(factures_labels),
        factures_values    = json.dumps(factures_values),

        # Dernières opérations
        dernieres_ops = dernieres_ops,
    )

# ================================================================
# FACTURES
# ================================================================
@app.route('/factures')
def liste_factures():
    if 'username' not in session:
        return redirect(url_for('login'))

    from sqlalchemy import extract
    filtre_etat    = request.args.get('etat', '')
    filtre_section = request.args.get('section', '')
    filtre_mois    = request.args.get('mois', '')
    filtre_client  = request.args.get('client', '')

    query = Facture.query
    if filtre_etat:    query = query.filter_by(etat_paiement=filtre_etat)
    if filtre_section: query = query.filter_by(section=filtre_section)
    if filtre_client:  query = query.filter(Facture.nom_client.ilike(f'%{filtre_client}%'))
    if filtre_mois:
        annee, mois = filtre_mois.split('-')
        query = query.filter(extract('year', Facture.date) == int(annee), extract('month', Facture.date) == int(mois))

    factures  = query.order_by(Facture.date.desc()).all()
    toutes    = Facture.query.all()
    total_ttc = sum(f.montant_ttc for f in factures)

    return render_template('factures.html',
        username=session['username'], role=session['role'],
        factures=factures, nb_factures=len(factures),
        nb_payees=sum(1 for f in toutes if f.etat_paiement == 'Payer'),
        nb_partielles=sum(1 for f in toutes if f.etat_paiement == 'Partiel'),
        nb_impayees=sum(1 for f in toutes if f.etat_paiement == 'Non Payer'),
        total_ttc=f"{total_ttc:,.0f}".replace(',', ' '),
        filtre_etat=filtre_etat, filtre_section=filtre_section,
        filtre_mois=filtre_mois, filtre_client=filtre_client,
    )


@app.route('/factures/nouvelle', methods=['GET', 'POST'])
def nouvelle_facture():
    if 'username' not in session:
        return redirect(url_for('login'))

    if request.method == 'POST':
        numero        = request.form['numero']
        nom_client    = request.form['nom_client']
        mode_paiement = request.form['mode_paiement']
        etat_paiement = request.form['etat_paiement']
        section       = request.form['section']
        tva_taux      = float(request.form.get('tva_taux', 18))

        services_f     = request.form.getlist('service[]')
        prix_unitaires = request.form.getlist('prix_unitaire[]')
        quantites      = request.form.getlist('quantite[]')

        montant_ttc_total = 0
        lignes_data = []
        for i in range(len(services_f)):
            svc = services_f[i].strip()
            if not svc: continue
            pu  = float(prix_unitaires[i]) if prix_unitaires[i] else 0
            qty = int(quantites[i]) if quantites[i] else 1
            ht  = pu * qty
            tva = round(ht * tva_taux / 100, 2)
            ttc = ht + tva
            montant_ttc_total += ttc
            lignes_data.append({'service': svc, 'prix_unitaire': pu, 'quantite': qty, 'montant_ht': ht, 'montant_ttc': ttc})
        if etat_paiement == 'Payer':
            montant_paye = montant_ttc_total
            reste_du     = 0
        elif etat_paiement == 'Partiel':
            montant_avance = float(request.form.get('montant_avance', 0) or 0)
            montant_paye   = montant_avance
            reste_du       = montant_ttc_total - montant_avance
        else:
            montant_paye = 0
            reste_du     = montant_ttc_total

        facture = Facture(
            numero=numero, nom_client=nom_client,
            service=lignes_data[0]['service'] if lignes_data else '',
            montant_ttc=montant_ttc_total, mode_paiement=mode_paiement,
            etat_paiement=etat_paiement, section=section,
            cree_par=session['username'], montant_paye=montant_paye, reste_du=reste_du,
        )
        db.session.add(facture)
        db.session.flush()

        for ld in lignes_data:
            db.session.add(LigneFacture(
                facture_id=facture.id, service=ld['service'],
                prix_unitaire=ld['prix_unitaire'], quantite=ld['quantite'],
                montant_ht=ld['montant_ht'], montant_ttc=ld['montant_ttc'],
            ))

        nb_op  = Operation.query.count() + 1
        op_num = f"OPE-{datetime.now().strftime('%d%m%Y')}-{nb_op:03d}"
        db.session.add(Operation(
            numero=op_num, nom_client=nom_client, service=facture.service,
            montant_ttc=montant_ttc_total, type_operation='Recettes',
            categorie='Facture', section=section, cree_par=session['username'],
        ))
        db.session.commit()
        flash(f'Facture {numero} enregistrée avec succès !', 'success')
        return redirect(url_for('liste_factures'))

    # GET
    numero_auto = generer_numero_facture('FKSP')

    # ✅ CORRECTION : services depuis la base de données
    services_list = Service.query.filter_by(actif=True).order_by(Service.section, Service.libelle).all()

    return render_template('nouvelle_facture.html',
        username=session['username'], role=session['role'],
        numero_auto=numero_auto, clients=Client.query.all(),
        today=datetime.now().strftime('%Y-%m-%d'),
        tva_defaut=18, services_list=services_list,
    )


@app.route('/factures/apercu/<int:id>')
def apercu_facture(id):
    if 'username' not in session:
        return redirect(url_for('login'))

    facture = Facture.query.get_or_404(id)
    client  = Client.query.filter_by(nom=facture.nom_client).first()
    params  = Parametres.query.first()
    lignes  = LigneFacture.query.filter_by(facture_id=facture.id).all()

    if not lignes:
        tva_taux   = 18
        montant_ht = round(facture.montant_ttc / (1 + tva_taux / 100), 2)
        lignes_display = [{'service': facture.service, 'quantite': 1, 'prix_unitaire': montant_ht, 'montant_ht': montant_ht, 'montant_ttc': facture.montant_ttc}]
    else:
        lignes_display = [{'service': l.service, 'quantite': l.quantite, 'prix_unitaire': l.prix_unitaire, 'montant_ht': l.montant_ht, 'montant_ttc': l.montant_ttc} for l in lignes]

    tva_taux  = 18
    total_ht  = sum(l['montant_ht'] for l in lignes_display)
    total_tva = round(facture.montant_ttc - total_ht, 2)

    return render_template('apercu_facture.html',
        facture=facture, client=client, params=params, lignes=lignes_display,
        total_ht=f"{total_ht:,.0f}".replace(',', ' '),
        total_tva=f"{total_tva:,.0f}".replace(',', ' '),
        montant_ttc=f"{facture.montant_ttc:,.0f}".replace(',', ' '),
        montant_paye=f"{facture.montant_paye:,.0f}".replace(',', ' '),
        reste_du=f"{facture.reste_du:,.0f}".replace(',', ' '),
        tva_taux=tva_taux, username=session['username'], role=session['role'],
    )


@app.route('/factures/modifier/<int:id>', methods=['POST'])
def modifier_facture(id):
    if 'username' not in session: return redirect(url_for('login'))
    if session['role'] == 'Lecture seule':
        flash('Accès refusé.', 'danger'); return redirect(url_for('liste_factures'))

    facture = Facture.query.get_or_404(id)
    facture.nom_client    = request.form.get('nom_client', '').strip()
    facture.service       = request.form.get('service', '').strip()
    facture.montant_ttc   = float(request.form.get('montant_ttc', 0))
    facture.mode_paiement = request.form.get('mode_paiement', '')
    facture.section       = request.form.get('section', '')
    nouvel_etat           = request.form.get('etat_paiement', '')

    if nouvel_etat == 'Payer':
        facture.montant_paye = facture.montant_ttc; facture.reste_du = 0
    elif nouvel_etat == 'Non Payer':
        facture.montant_paye = 0; facture.reste_du = facture.montant_ttc

    facture.etat_paiement = nouvel_etat
    db.session.commit()
    flash(f'Facture {facture.numero} modifiée avec succès !', 'success')
    return redirect(url_for('liste_factures'))


@app.route('/factures/supprimer/<int:id>', methods=['POST'])
def supprimer_facture(id):
    if 'username' not in session: return redirect(url_for('login'))
    if session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger'); return redirect(url_for('liste_factures'))

    facture = Facture.query.get_or_404(id)
    numero  = facture.numero
    LigneFacture.query.filter_by(facture_id=id).delete()
    Paiement.query.filter_by(facture_id=id).delete()
    db.session.delete(facture)
    db.session.commit()
    flash(f'Facture {numero} supprimée.', 'warning')
    return redirect(url_for('liste_factures'))


@app.route('/factures/paiement/<int:id>', methods=['GET', 'POST'])
def enregistrer_paiement(id):
    if 'username' not in session: return redirect(url_for('login'))
    if session['role'] == 'Lecture seule':
        flash('Accès refusé.', 'danger'); return redirect(url_for('liste_factures'))

    facture = Facture.query.get_or_404(id)

    if request.method == 'POST':
        montant_paye_new = float(request.form.get('montant_paye', 0))
        mode_paiement    = request.form.get('mode_paiement', 'Espece')
        notes            = request.form.get('notes', '').strip()

        if montant_paye_new <= 0:
            flash('Le montant payé doit être supérieur à 0.', 'danger')
            return redirect(url_for('enregistrer_paiement', id=id))
        if montant_paye_new > facture.reste_du:
            flash(f'Le montant dépasse le reste dû ({facture.reste_du:,.0f} FCFA).', 'danger')
            return redirect(url_for('enregistrer_paiement', id=id))

        facture.montant_paye += montant_paye_new
        facture.reste_du     -= montant_paye_new
        if facture.reste_du <= 0:
            facture.etat_paiement = 'Payer'; facture.reste_du = 0
        else:
            facture.etat_paiement = 'Partiel'

        nb  = Paiement.query.count() + 1
        num = f"PAI-{datetime.now().strftime('%d%m%Y')}-{nb:03d}"
        db.session.add(Paiement(
            numero=num, facture_id=facture.id, n_facture=facture.numero,
            nom_client=facture.nom_client, montant_facture=facture.montant_ttc,
            montant_paye=montant_paye_new, reste_du=facture.reste_du,
            mode_paiement=mode_paiement, etat_facture=facture.etat_paiement,
            notes=notes, cree_par=session['username'],
        ))
        db.session.commit()
        flash(f'Paiement de {montant_paye_new:,.0f} FCFA enregistré !', 'success')
        return redirect(url_for('liste_factures'))

    return render_template('enregistrer_paiement.html',
        facture=facture, username=session['username'], role=session['role'],
        montant_paye=f"{facture.montant_paye:,.0f}".replace(',', ' '),
        reste_du=f"{facture.reste_du:,.0f}".replace(',', ' '),
        montant_ttc=f"{facture.montant_ttc:,.0f}".replace(',', ' '),
    )


@app.route('/paiements')
def historique_paiements():
    if 'username' not in session: return redirect(url_for('login'))
    paiements      = Paiement.query.order_by(Paiement.date.desc()).all()
    total_encaisse = sum(p.montant_paye for p in paiements)
    return render_template('historique_paiements.html',
        username=session['username'], role=session['role'],
        paiements=paiements, nb_paiements=len(paiements),
        total_encaisse=f"{total_encaisse:,.0f}".replace(',', ' '),
    )

# ================================================================
# OPÉRATIONS
# ================================================================
@app.route('/operations')
def liste_operations():
    if 'username' not in session: return redirect(url_for('login'))

    from sqlalchemy import extract
    filtre_type    = request.args.get('type', '')
    filtre_section = request.args.get('section', '')
    filtre_mois    = request.args.get('mois', '')

    query = Operation.query
    if filtre_type:    query = query.filter_by(type_operation=filtre_type)
    if filtre_section: query = query.filter_by(section=filtre_section)
    if filtre_mois:
        annee, mois = filtre_mois.split('-')
        query = query.filter(extract('year', Operation.date) == int(annee), extract('month', Operation.date) == int(mois))

    operations     = query.order_by(Operation.date.desc()).all()
    total_recettes = sum(o.montant_ttc for o in operations if o.type_operation == 'Recettes')
    total_depenses = sum(o.montant_ttc for o in operations if o.type_operation == 'Depenses')
    solde          = total_recettes - total_depenses

    return render_template('operations.html',
        username=session['username'], role=session['role'],
        operations=operations, nb_operations=len(operations),
        total_recettes=f"{total_recettes:,.0f}".replace(',', ' '),
        total_depenses=f"{total_depenses:,.0f}".replace(',', ' '),
        solde=f"{solde:,.0f}".replace(',', ' '), solde_positif=(solde >= 0),
        filtre_type=filtre_type, filtre_section=filtre_section, filtre_mois=filtre_mois,
    )


@app.route('/operations/nouvelle', methods=['GET', 'POST'])
def nouvelle_operation():
    if 'username' not in session: return redirect(url_for('login'))
    if session['role'] == 'Lecture seule':
        flash('Accès refusé.', 'danger'); return redirect(url_for('liste_operations'))

    if request.method == 'POST':
        nb     = Operation.query.count() + 1
        numero = f"OPE-{datetime.now().strftime('%d%m%Y')}-{nb:03d}"
        db.session.add(Operation(
            numero=numero, nom_client=request.form.get('nom_client', ''),
            service=request.form['service'], montant_ttc=float(request.form['montant_ttc']),
            type_operation=request.form['type_operation'],
            categorie=request.form.get('categorie', 'Manuel'),
            section=request.form['section'], cree_par=session['username'],
        ))
        db.session.commit()
        flash(f'Opération {numero} enregistrée !', 'success')
        return redirect(url_for('liste_operations'))

    return render_template('nouvelle_operation.html',
        username=session['username'], role=session['role'],
        today=datetime.now().strftime('%Y-%m-%d'), clients=Client.query.all(),
    )


@app.route('/operations/supprimer/<int:id>', methods=['POST'])
def supprimer_operation(id):
    if 'username' not in session: return redirect(url_for('login'))
    if session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger'); return redirect(url_for('liste_operations'))
    op = Operation.query.get_or_404(id)
    db.session.delete(op)
    db.session.commit()
    flash(f'Opération {op.numero} supprimée.', 'warning')
    return redirect(url_for('liste_operations'))

# ================================================================
# CLIENTS
# ================================================================
@app.route('/clients')
def liste_clients():
    if 'username' not in session: return redirect(url_for('login'))
    recherche = request.args.get('q', '')
    if recherche:
        clients = Client.query.filter(Client.nom.ilike(f'%{recherche}%') | Client.telephone.ilike(f'%{recherche}%') | Client.numero.ilike(f'%{recherche}%')).order_by(Client.nom).all()
    else:
        clients = Client.query.order_by(Client.nom).all()
    return render_template('clients.html', username=session['username'], role=session['role'], clients=clients, nb_clients=len(clients), recherche=recherche)


@app.route('/clients/nouveau', methods=['GET', 'POST'])
def nouveau_client():
    if 'username' not in session: return redirect(url_for('login'))
    if session['role'] == 'Lecture seule':
        flash('Accès refusé.', 'danger'); return redirect(url_for('liste_clients'))
    if request.method == 'POST':
        nb = Client.query.count() + 1
        numero = f"CLI-{nb:03d}"
        while Client.query.filter_by(numero=numero).first():
            nb += 1; numero = f"CLI-{nb:03d}"
        db.session.add(Client(
            numero=numero, nom=request.form['nom'].strip().upper(),
            adresse=request.form.get('adresse','').strip(),
            telephone=request.form.get('telephone','').strip(),
            email=request.form.get('email','').strip(),
            nif=request.form.get('nif','').strip(),
            rccm=request.form.get('rccm','').strip(),
        ))
        db.session.commit()
        flash(f'Client enregistré ({numero}) !', 'success')
        return redirect(url_for('liste_clients'))
    return render_template('nouveau_client.html', username=session['username'], role=session['role'])


@app.route('/clients/modifier/<int:id>', methods=['GET', 'POST'])
def modifier_client(id):
    if 'username' not in session: return redirect(url_for('login'))
    if session['role'] == 'Lecture seule':
        flash('Accès refusé.', 'danger'); return redirect(url_for('liste_clients'))
    client = Client.query.get_or_404(id)
    if request.method == 'POST':
        client.nom = request.form['nom'].strip().upper()
        client.adresse = request.form.get('adresse','').strip()
        client.telephone = request.form.get('telephone','').strip()
        client.email = request.form.get('email','').strip()
        client.nif = request.form.get('nif','').strip()
        client.rccm = request.form.get('rccm','').strip()
        db.session.commit()
        flash(f'Client {client.nom} modifié !', 'success')
        return redirect(url_for('liste_clients'))
    return render_template('nouveau_client.html', username=session['username'], role=session['role'], client=client, modification=True)


@app.route('/clients/supprimer/<int:id>', methods=['POST'])
def supprimer_client(id):
    if 'username' not in session: return redirect(url_for('login'))
    if session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger'); return redirect(url_for('liste_clients'))
    client = Client.query.get_or_404(id)
    nom = client.nom
    db.session.delete(client); db.session.commit()
    flash(f'Client {nom} supprimé.', 'warning')
    return redirect(url_for('liste_clients'))

# ================================================================
# SERVICES
# ================================================================
@app.route('/services')
def liste_services():
    if 'username' not in session: return redirect(url_for('login'))
    if session['role'] not in ['Administrateur', 'Caissier']:
        flash('Accès refusé.', 'danger'); return redirect(url_for('dashboard'))
    return render_template('services.html',
        username=session['username'], role=session['role'],
        services_studio=Service.query.filter_by(section='Studio', actif=True).order_by(Service.libelle).all(),
        services_sono=Service.query.filter_by(section='Sonorisation', actif=True).order_by(Service.libelle).all(),
        services_autres=Service.query.filter(Service.section.notin_(['Studio','Sonorisation']), Service.actif==True).order_by(Service.libelle).all(),
        total=Service.query.filter_by(actif=True).count(),
    )


@app.route('/services/nouveau', methods=['POST'])
def nouveau_service():
    if 'username' not in session: return redirect(url_for('login'))
    if session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger'); return redirect(url_for('liste_services'))
    libelle = request.form.get('libelle','').strip()
    section = request.form.get('section','').strip()
    prix    = float(request.form.get('prix', 0) or 0)
    if not libelle or not section:
        flash('Libellé et section obligatoires.', 'danger'); return redirect(url_for('liste_services'))
    if Service.query.filter_by(libelle=libelle, section=section).first():
        flash(f'Service "{libelle}" existe déjà.', 'warning'); return redirect(url_for('liste_services'))
    db.session.add(Service(section=section, libelle=libelle, prix=prix))
    db.session.commit()
    flash(f'Service "{libelle}" ajouté !', 'success')
    return redirect(url_for('liste_services'))


@app.route('/services/modifier/<int:id>', methods=['POST'])
def modifier_service(id):
    if 'username' not in session: return redirect(url_for('login'))
    if session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger'); return redirect(url_for('liste_services'))
    svc = Service.query.get_or_404(id)
    svc.libelle = request.form.get('libelle','').strip()
    svc.section = request.form.get('section','').strip()
    svc.prix    = float(request.form.get('prix', 0) or 0)
    db.session.commit()
    flash(f'Service "{svc.libelle}" modifié !', 'success')
    return redirect(url_for('liste_services'))


@app.route('/services/supprimer/<int:id>', methods=['POST'])
def supprimer_service(id):
    if 'username' not in session: return redirect(url_for('login'))
    if session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger'); return redirect(url_for('liste_services'))
    svc = Service.query.get_or_404(id)
    svc.actif = False
    db.session.commit()
    flash(f'Service "{svc.libelle}" supprimé.', 'warning')
    return redirect(url_for('liste_services'))


@app.route('/api/services')
def api_services():
    services = Service.query.filter_by(actif=True).order_by(Service.section, Service.libelle).all()
    return jsonify([{'id': s.id, 'section': s.section, 'libelle': s.libelle, 'prix': s.prix} for s in services])


@app.route('/services/fiche-tarifs')
def fiche_tarifs():
    if 'username' not in session: return redirect(url_for('login'))
    params = Parametres.query.first()
    return render_template('fiche_tarifs.html',
        params=params, username=session['username'], role=session['role'],
        services_studio=Service.query.filter_by(section='Studio', actif=True).order_by(Service.libelle).all(),
        services_sono=Service.query.filter_by(section='Sonorisation', actif=True).order_by(Service.libelle).all(),
        services_autres=Service.query.filter(Service.section.notin_(['Studio','Sonorisation']), Service.actif==True).order_by(Service.libelle).all(),
        date_maj=datetime.now().strftime('%d/%m/%Y'), annee=datetime.now().year,
    )

# ================================================================
# UTILISATEURS
# ================================================================
@app.route('/utilisateurs')
def liste_utilisateurs():
    if 'username' not in session: return redirect(url_for('login'))
    if session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger'); return redirect(url_for('dashboard'))
    utilisateurs = Utilisateur.query.order_by(Utilisateur.role, Utilisateur.username).all()
    return render_template('utilisateurs.html',
        username=session['username'], role=session['role'],
        utilisateurs=utilisateurs, nb_total=len(utilisateurs),
        nb_actifs=sum(1 for u in utilisateurs if u.statut == 'Actif'),
    )


@app.route('/utilisateurs/nouveau', methods=['POST'])
def nouveau_utilisateur():
    if 'username' not in session or session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger'); return redirect(url_for('liste_utilisateurs'))
    username    = request.form.get('username','').strip().lower()
    password    = request.form.get('password','').strip()
    role        = request.form.get('role','').strip()
    nom_complet = request.form.get('nom_complet','').strip()
    if not username or not password or not role:
        flash('Tous les champs sont obligatoires.', 'danger'); return redirect(url_for('liste_utilisateurs'))
    if Utilisateur.query.filter_by(username=username).first():
        flash(f'Utilisateur "{username}" existe déjà.', 'warning'); return redirect(url_for('liste_utilisateurs'))
    db.session.add(Utilisateur(username=username, password=generate_password_hash(password), role=role, nom_complet=nom_complet or username, statut='Actif'))
    db.session.commit()
    flash(f'Utilisateur "{username}" créé !', 'success')
    return redirect(url_for('liste_utilisateurs'))


@app.route('/utilisateurs/modifier/<int:id>', methods=['POST'])
def modifier_utilisateur(id):
    if 'username' not in session or session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger'); return redirect(url_for('liste_utilisateurs'))
    user = Utilisateur.query.get_or_404(id)
    nouveau_role = request.form.get('role','').strip()
    if user.username == session['username'] and nouveau_role != 'Administrateur':
        flash('Vous ne pouvez pas changer votre propre rôle.', 'warning'); return redirect(url_for('liste_utilisateurs'))
    user.nom_complet = request.form.get('nom_complet','').strip()
    user.role        = nouveau_role
    user.statut      = request.form.get('statut','Actif').strip()
    nouveau_mdp = request.form.get('password','').strip()
    if nouveau_mdp: user.password = generate_password_hash(nouveau_mdp)
    db.session.commit()
    flash(f'Utilisateur "{user.username}" modifié !', 'success')
    return redirect(url_for('liste_utilisateurs'))


@app.route('/utilisateurs/toggle/<int:id>', methods=['POST'])
def toggle_utilisateur(id):
    if 'username' not in session or session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger'); return redirect(url_for('liste_utilisateurs'))
    user = Utilisateur.query.get_or_404(id)
    if user.username == session['username']:
        flash('Vous ne pouvez pas désactiver votre compte.', 'warning'); return redirect(url_for('liste_utilisateurs'))
    user.statut = 'Inactif' if user.statut == 'Actif' else 'Actif'
    db.session.commit()
    flash(f'Compte "{user.username}" {"désactivé" if user.statut == "Inactif" else "activé"}.', 'success')
    return redirect(url_for('liste_utilisateurs'))


@app.route('/utilisateurs/supprimer/<int:id>', methods=['POST'])
def supprimer_utilisateur(id):
    if 'username' not in session or session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger'); return redirect(url_for('liste_utilisateurs'))
    user = Utilisateur.query.get_or_404(id)
    if user.username == session['username']:
        flash('Vous ne pouvez pas supprimer votre compte.', 'danger'); return redirect(url_for('liste_utilisateurs'))
    db.session.delete(user); db.session.commit()
    flash(f'Utilisateur "{user.username}" supprimé.', 'warning')
    return redirect(url_for('liste_utilisateurs'))

# ================================================================
# PARAMÈTRES
# ================================================================
@app.route('/parametres', methods=['GET', 'POST'])
def parametres():
    if 'username' not in session: return redirect(url_for('login'))
    if session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger'); return redirect(url_for('dashboard'))

    params = Parametres.query.first()
    if not params:
        params = Parametres(); db.session.add(params); db.session.commit()

    if request.method == 'POST':
        params.nom_entreprise      = request.form.get('nom_entreprise','').strip()
        params.slogan              = request.form.get('slogan','').strip()
        params.adresse             = request.form.get('adresse','').strip()
        params.telephone           = request.form.get('telephone','').strip()
        params.email               = request.form.get('email','').strip()
        params.site_web            = request.form.get('site_web','').strip()
        params.nif                 = request.form.get('nif','').strip()
        params.rccm                = request.form.get('rccm','').strip()
        params.couleur_principale  = request.form.get('couleur_principale','#e94560').strip()
        params.mentions_legales    = request.form.get('mentions_legales','').strip()
        params.coordonnees_bancaires = request.form.get('coordonnees_bancaires','').strip()

        if request.form.get('supprimer_logo') == '1' and params.logo_filename:
            chemin = os.path.join(app.config['UPLOAD_FOLDER'], params.logo_filename)
            if os.path.exists(chemin): os.remove(chemin)
            params.logo_filename = None

        logo = request.files.get('logo')
        if logo and logo.filename:
            filename = secure_filename(logo.filename)
            logo.save(os.path.join(app.config['UPLOAD_FOLDER'], filename))
            params.logo_filename = filename

        db.session.commit()
        flash('Paramètres enregistrés !', 'success')
        return redirect(url_for('parametres'))

    return render_template('parametres.html', params=params, username=session['username'], role=session['role'])

# ================================================================
# RAPPORT
# ================================================================
@app.route('/rapport', methods=['GET'])
def rapport():
    if 'username' not in session: return redirect(url_for('login'))

    from datetime import date
    today      = date.today()
    date_debut = request.args.get('date_debut', today.replace(day=1).strftime('%Y-%m-%d'))
    date_fin   = request.args.get('date_fin',   today.strftime('%Y-%m-%d'))
    section    = request.args.get('section', '')

    try:
        d_debut = datetime.strptime(date_debut, '%Y-%m-%d')
        d_fin   = datetime.strptime(date_fin,   '%Y-%m-%d').replace(hour=23, minute=59, second=59)
    except:
        d_debut = datetime(today.year, today.month, 1); d_fin = datetime.now()

    query_ops = Operation.query.filter(Operation.date >= d_debut, Operation.date <= d_fin)
    if section: query_ops = query_ops.filter_by(section=section)
    operations = query_ops.order_by(Operation.date.desc()).all()

    total_recettes = sum(o.montant_ttc for o in operations if o.type_operation == 'Recettes')
    total_depenses = sum(o.montant_ttc for o in operations if o.type_operation == 'Depenses')
    solde_net      = total_recettes - total_depenses

    query_fac = Facture.query.filter(Facture.date >= d_debut, Facture.date <= d_fin)
    if section: query_fac = query_fac.filter_by(section=section)
    factures = query_fac.all()

    sections_data, clients_data, depenses_data = {}, {}, {}
    for op in operations:
        if op.type_operation == 'Recettes':
            s = op.section or 'Autre'; sections_data[s] = sections_data.get(s, 0) + op.montant_ttc
            c = op.nom_client or 'Inconnu'; clients_data[c] = clients_data.get(c, 0) + op.montant_ttc
        else:
            cat = op.service or 'Autre'; depenses_data[cat] = depenses_data.get(cat, 0) + op.montant_ttc

    top_clients = sorted(clients_data.items(), key=lambda x: x[1], reverse=True)[:5]

    return render_template('rapport.html',
        username=session['username'], role=session['role'],
        date_debut=date_debut, date_fin=date_fin, section=section,
        total_recettes=f"{total_recettes:,.0f}".replace(',', ' '),
        total_depenses=f"{total_depenses:,.0f}".replace(',', ' '),
        solde_net=f"{solde_net:,.0f}".replace(',', ' '), solde_positif=solde_net >= 0,
        nb_operations=len(operations), operations=operations,
        nb_factures=len(factures),
        nb_payees=sum(1 for f in factures if f.etat_paiement == 'Payer'),
        nb_partielles=sum(1 for f in factures if f.etat_paiement == 'Partiel'),
        nb_impayees=sum(1 for f in factures if f.etat_paiement == 'Non Payer'),
        total_facture=f"{sum(f.montant_ttc for f in factures):,.0f}".replace(',', ' '),
        total_encaisse=f"{sum(f.montant_paye for f in factures):,.0f}".replace(',', ' '),
        total_impayes=f"{sum(f.reste_du for f in factures):,.0f}".replace(',', ' '),
        sections_labels=json.dumps(list(sections_data.keys())),
        sections_values=json.dumps(list(sections_data.values())),
        top_clients=top_clients, depenses_data=depenses_data,
    )

# ================================================================
# EXPORT EXCEL
# ================================================================
@app.route('/export/factures')
def export_factures():
    if 'username' not in session: return redirect(url_for('login'))
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Factures"
    ROUGE = "C0392B"; BLANC = "FFFFFF"; GRIS = "F5F5F5"; VERT = "27AE60"; ORANGE = "E67E22"

    ws.merge_cells('A1:J1')
    ws['A1'] = f"KS Production — Export Factures — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A1'].font = Font(bold=True, size=12, color=BLANC)
    ws['A1'].fill = PatternFill("solid", fgColor=ROUGE)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    headers = ['N° Facture','Date','Client','Service','Section','Mode Paiement','État','Montant TTC','Montant Payé','Reste Dû']
    for col, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color=BLANC, size=10)
        cell.fill = PatternFill("solid", fgColor="1A1A2E")
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    factures = Facture.query.order_by(Facture.date.desc()).all()
    for i, f in enumerate(factures):
        row = i + 3; fill = GRIS if i % 2 == 0 else BLANC
        etat = 'Payée' if f.etat_paiement == 'Payer' else ('Partielle' if f.etat_paiement == 'Partiel' else 'Impayée')
        data = [f.numero, f.date.strftime('%d/%m/%Y') if f.date else '', f.nom_client or '', f.service or '', f.section or '', f.mode_paiement or '', etat, round(f.montant_ttc,0), round(f.montant_paye,0), round(f.reste_du,0)]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=fill)
            cell.alignment = Alignment(vertical='center')
            cell.font = Font(size=10)
            if col in [8,9,10]: cell.number_format = '#,##0'; cell.alignment = Alignment(horizontal='right', vertical='center')
            if col == 7: cell.font = Font(color=VERT if val=='Payée' else (ORANGE if val=='Partielle' else ROUGE), bold=True, size=10)

    total_row = len(factures) + 3
    ws.cell(row=total_row, column=7, value="TOTAL").font = Font(bold=True, size=10)
    for col, field in [(8,'montant_ttc'),(9,'montant_paye'),(10,'reste_du')]:
        cell = ws.cell(row=total_row, column=col, value=round(sum(getattr(f,field) for f in factures),0))
        cell.font = Font(bold=True, size=10, color=BLANC); cell.fill = PatternFill("solid", fgColor=ROUGE)
        cell.number_format = '#,##0'; cell.alignment = Alignment(horizontal='right', vertical='center')

    for col, w in enumerate([18,12,22,30,14,16,12,16,16,14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"KS_Factures_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")


@app.route('/export/operations')
def export_operations():
    if 'username' not in session: return redirect(url_for('login'))
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Opérations"
    ROUGE="C0392B"; BLANC="FFFFFF"; GRIS="F5F5F5"; VERT="27AE60"; DARK="1A1A2E"

    ws.merge_cells('A1:H1')
    ws['A1'] = f"KS Production — Export Opérations — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A1'].font = Font(bold=True, size=12, color=BLANC)
    ws['A1'].fill = PatternFill("solid", fgColor=ROUGE)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    for col, h in enumerate(['N° Opération','Date','Client','Service','Section','Type','Catégorie','Montant TTC'], 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color=BLANC, size=10); cell.fill = PatternFill("solid", fgColor=DARK)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    ops = Operation.query.order_by(Operation.date.desc()).all()
    for i, op in enumerate(ops):
        row = i + 3; fill = GRIS if i % 2 == 0 else BLANC
        data = [op.numero, op.date.strftime('%d/%m/%Y') if op.date else '', op.nom_client or '', op.service or '', op.section or '', op.type_operation or '', op.categorie or '', round(op.montant_ttc,0)]
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=fill); cell.font = Font(size=10); cell.alignment = Alignment(vertical='center')
            if col == 8: cell.number_format = '#,##0'; cell.alignment = Alignment(horizontal='right', vertical='center')
            if col == 6: cell.font = Font(size=10, color=VERT if val=='Recettes' else ROUGE, bold=True)

    total_row = len(ops) + 3
    for label, val, color in [("Recettes", sum(o.montant_ttc for o in ops if o.type_operation=='Recettes'), VERT), ("Dépenses", sum(o.montant_ttc for o in ops if o.type_operation=='Depenses'), ROUGE)]:
        ws.cell(row=total_row, column=6, value="TOTAL")
        cell = ws.cell(row=total_row, column=8, value=round(val,0))
        cell.font = Font(bold=True, size=10, color=BLANC); cell.fill = PatternFill("solid", fgColor=color)
        cell.number_format = '#,##0'; total_row += 1

    for col, w in enumerate([18,12,22,30,14,12,14,16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"KS_Operations_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")


@app.route('/export/paiements')
def export_paiements():
    if 'username' not in session: return redirect(url_for('login'))
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Paiements"
    ROUGE="C0392B"; BLANC="FFFFFF"; GRIS="F5F5F5"; DARK="1A1A2E"

    ws.merge_cells('A1:I1')
    ws['A1'] = f"KS Production — Export Paiements — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A1'].font = Font(bold=True, size=12, color=BLANC); ws['A1'].fill = PatternFill("solid", fgColor=ROUGE)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center'); ws.row_dimensions[1].height = 28

    for col, h in enumerate(['N° Paiement','Date','N° Facture','Client','Mode','Montant Encaissé','Reste Dû','Statut','Saisi par'], 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color=BLANC, size=10); cell.fill = PatternFill("solid", fgColor=DARK)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    paiements = Paiement.query.order_by(Paiement.date.desc()).all()
    for i, p in enumerate(paiements):
        row = i + 3; fill = GRIS if i % 2 == 0 else BLANC
        data = [p.numero, p.date.strftime('%d/%m/%Y') if p.date else '', p.n_facture or '', p.nom_client or '', p.mode_paiement or '', round(p.montant_paye,0), round(p.reste_du,0), 'Payée' if p.etat_facture=='Payer' else 'Partielle', p.cree_par or '']
        for col, val in enumerate(data, 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=fill); cell.font = Font(size=10); cell.alignment = Alignment(vertical='center')
            if col in [6,7]: cell.number_format = '#,##0'; cell.alignment = Alignment(horizontal='right', vertical='center')

    total_row = len(paiements) + 3
    ws.cell(row=total_row, column=5, value="TOTAL ENCAISSÉ").font = Font(bold=True, size=10)
    cell = ws.cell(row=total_row, column=6, value=round(sum(p.montant_paye for p in paiements),0))
    cell.font = Font(bold=True, color=BLANC, size=10); cell.fill = PatternFill("solid", fgColor="27AE60"); cell.number_format = '#,##0'

    for col, w in enumerate([20,12,18,22,16,18,14,12,14], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"KS_Paiements_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")


@app.route('/export/clients')
def export_clients():
    if 'username' not in session: return redirect(url_for('login'))
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils import get_column_letter

    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Clients"
    ROUGE="C0392B"; BLANC="FFFFFF"; GRIS="F5F5F5"; DARK="1A1A2E"

    ws.merge_cells('A1:G1')
    ws['A1'] = f"KS Production — Export Clients — {datetime.now().strftime('%d/%m/%Y %H:%M')}"
    ws['A1'].font = Font(bold=True, size=12, color=BLANC); ws['A1'].fill = PatternFill("solid", fgColor=ROUGE)
    ws['A1'].alignment = Alignment(horizontal='center', vertical='center'); ws.row_dimensions[1].height = 28

    for col, h in enumerate(['N° Client','Nom','Téléphone','Email','Adresse','NIF','RCCM'], 1):
        cell = ws.cell(row=2, column=col, value=h)
        cell.font = Font(bold=True, color=BLANC, size=10); cell.fill = PatternFill("solid", fgColor=DARK)
        cell.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[2].height = 22

    clients = Client.query.order_by(Client.nom).all()
    for i, c in enumerate(clients):
        row = i + 3; fill = GRIS if i % 2 == 0 else BLANC
        for col, val in enumerate([c.numero, c.nom, c.telephone or '', c.email or '', c.adresse or '', c.nif or '', c.rccm or ''], 1):
            cell = ws.cell(row=row, column=col, value=val)
            cell.fill = PatternFill("solid", fgColor=fill); cell.font = Font(size=10); cell.alignment = Alignment(vertical='center')

    for col, w in enumerate([14,28,16,28,28,16,16], 1):
        ws.column_dimensions[get_column_letter(col)].width = w

    buf = io.BytesIO(); wb.save(buf); buf.seek(0)
    return send_file(buf, mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
                     as_attachment=True, download_name=f"KS_Clients_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx")

# ================================================================

# ================================================================
# ROUTE RÉINITIALISATION — à ajouter dans app.py
# Réservé exclusivement à l'Administrateur
# ================================================================

@app.route('/reinitialiser', methods=['GET', 'POST'])
def reinitialiser():
    if 'username' not in session:
        return redirect(url_for('login'))
    if session['role'] != 'Administrateur':
        flash('Accès refusé : réservé aux Administrateurs.', 'danger')
        return redirect(url_for('dashboard'))

    if request.method == 'POST':
        confirmation = request.form.get('confirmation', '').strip()

        # Vérifier le mot de confirmation
        if confirmation != 'REINITIALISER':
            flash('Mot de confirmation incorrect. Tapez exactement : REINITIALISER', 'danger')
            return redirect(url_for('reinitialiser'))

        options = request.form.getlist('options')

        try:
            if 'factures' in options:
                LigneFacture.query.delete()
                Paiement.query.delete()
                Facture.query.delete()
                db.session.commit()

            if 'operations' in options:
                Operation.query.delete()
                db.session.commit()

            if 'clients' in options:
                Client.query.delete()
                db.session.commit()

            if 'services' in options:
                Service.query.delete()
                db.session.commit()

            flash('✅ Réinitialisation effectuée avec succès ! L\'application est prête pour une utilisation réelle.', 'success')

        except Exception as e:
            db.session.rollback()
            flash(f'Erreur lors de la réinitialisation : {str(e)}', 'danger')

        return redirect(url_for('dashboard'))

    # Stats actuelles pour afficher ce qui sera supprimé
    return render_template('reinitialiser.html',
        username      = session['username'],
        role          = session['role'],
        nb_factures   = Facture.query.count(),
        nb_operations = Operation.query.count(),
        nb_clients    = Client.query.count(),
        nb_services   = Service.query.filter_by(actif=True).count(),
        nb_paiements  = Paiement.query.count(),
    )


# ================================================================
# ROUTES SAUVEGARDE & RESTAURATION — à ajouter dans app.py
# ================================================================

@app.route('/sauvegarde')
def sauvegarde():
    if 'username' not in session:
        return redirect(url_for('login'))
    if session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger')
        return redirect(url_for('dashboard'))

    return render_template('sauvegarde.html',
        username      = session['username'],
        role          = session['role'],
        nb_factures   = Facture.query.count(),
        nb_operations = Operation.query.count(),
        nb_clients    = Client.query.count(),
        nb_services   = Service.query.count(),
        nb_paiements  = Paiement.query.count(),
        nb_params     = Parametres.query.count(),
    )


@app.route('/sauvegarde/exporter')
def exporter_donnees():
    if 'username' not in session:
        return redirect(url_for('login'))
    if session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger')
        return redirect(url_for('dashboard'))

    # ── Collecter toutes les données ──
    data = {
        'meta': {
            'version'    : '1.0',
            'application': 'KS Production Web',
            'date_export': datetime.now().strftime('%d/%m/%Y %H:%M:%S'),
            'exporte_par': session['username'],
        },
        'parametres': [],
        'clients'   : [],
        'services'  : [],
        'factures'  : [],
        'lignes'    : [],
        'operations': [],
        'paiements' : [],
    }

    # Paramètres
    params = Parametres.query.first()
    if params:
        data['parametres'].append({
            'nom_entreprise'     : params.nom_entreprise,
            'slogan'             : params.slogan,
            'adresse'            : params.adresse,
            'telephone'          : params.telephone,
            'email'              : params.email,
            'site_web'           : params.site_web,
            'nif'                : params.nif,
            'rccm'               : params.rccm,
            'couleur_principale' : params.couleur_principale,
            'mentions_legales'   : params.mentions_legales,
            'coordonnees_bancaires': params.coordonnees_bancaires,
            'logo_filename'      : params.logo_filename,
        })

    # Clients
    for c in Client.query.all():
        data['clients'].append({
            'numero'   : c.numero,
            'nom'      : c.nom,
            'adresse'  : c.adresse,
            'telephone': c.telephone,
            'email'    : c.email,
            'nif'      : c.nif,
            'rccm'     : c.rccm,
        })

    # Services
    for s in Service.query.all():
        data['services'].append({
            'section': s.section,
            'libelle': s.libelle,
            'prix'   : s.prix,
            'actif'  : s.actif,
        })

    # Factures
    for f in Facture.query.all():
        data['factures'].append({
            'numero'        : f.numero,
            'date'          : f.date.strftime('%Y-%m-%d') if f.date else None,
            'nom_client'    : f.nom_client,
            'service'       : f.service,
            'montant_ttc'   : f.montant_ttc,
            'mode_paiement' : f.mode_paiement,
            'etat_paiement' : f.etat_paiement,
            'section'       : f.section,
            'cree_par'      : f.cree_par,
            'montant_paye'  : f.montant_paye,
            'reste_du'      : f.reste_du,
        })

    # Lignes de facture
    for l in LigneFacture.query.all():
        # Récupérer le numéro de facture correspondant
        fac = Facture.query.get(l.facture_id)
        data['lignes'].append({
            'facture_numero': fac.numero if fac else None,
            'service'       : l.service,
            'prix_unitaire' : l.prix_unitaire,
            'quantite'      : l.quantite,
            'montant_ht'    : l.montant_ht,
            'montant_ttc'   : l.montant_ttc,
        })

    # Opérations
    for o in Operation.query.all():
        data['operations'].append({
            'numero'        : o.numero,
            'date'          : o.date.strftime('%Y-%m-%d') if o.date else None,
            'nom_client'    : o.nom_client,
            'service'       : o.service,
            'montant_ttc'   : o.montant_ttc,
            'type_operation': o.type_operation,
            'categorie'     : o.categorie,
            'section'       : o.section,
            'cree_par'      : o.cree_par,
        })

    # Paiements
    for p in Paiement.query.all():
        fac = Facture.query.get(p.facture_id)
        data['paiements'].append({
            'numero'         : p.numero,
            'date'           : p.date.strftime('%Y-%m-%d') if p.date else None,
            'facture_numero' : fac.numero if fac else p.n_facture,
            'n_facture'      : p.n_facture,
            'nom_client'     : p.nom_client,
            'montant_facture': p.montant_facture,
            'montant_paye'   : p.montant_paye,
            'reste_du'       : p.reste_du,
            'mode_paiement'  : p.mode_paiement,
            'etat_facture'   : p.etat_facture,
            'notes'          : p.notes,
            'cree_par'       : p.cree_par,
        })

    # ── Envoyer le fichier JSON ──
    import json as json_module
    contenu = json_module.dumps(data, ensure_ascii=False, indent=2)
    nom_fichier = f"KS_Backup_{datetime.now().strftime('%d%m%Y_%H%M')}.json"

    buf = io.BytesIO(contenu.encode('utf-8'))
    buf.seek(0)
    return send_file(buf, mimetype='application/json',
                     as_attachment=True, download_name=nom_fichier)


@app.route('/sauvegarde/importer', methods=['POST'])
def importer_donnees():
    if 'username' not in session:
        return redirect(url_for('login'))
    if session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger')
        return redirect(url_for('sauvegarde'))

    import json as json_module

    fichier = request.files.get('fichier_backup')
    mode    = request.form.get('mode', 'fusionner')  # 'fusionner' ou 'remplacer'

    if not fichier or fichier.filename == '':
        flash('Aucun fichier sélectionné.', 'danger')
        return redirect(url_for('sauvegarde'))

    if not fichier.filename.endswith('.json'):
        flash('Format invalide. Veuillez importer un fichier .json', 'danger')
        return redirect(url_for('sauvegarde'))

    try:
        contenu = fichier.read().decode('utf-8')
        data    = json_module.loads(contenu)

        # Vérifier que c'est un backup valide
        if 'meta' not in data or data['meta'].get('application') != 'KS Production Web':
            flash('Fichier invalide. Ce n\'est pas un backup KS Production.', 'danger')
            return redirect(url_for('sauvegarde'))

        # ── Mode REMPLACER : vider les tables d'abord ──
        if mode == 'remplacer':
            LigneFacture.query.delete()
            Paiement.query.delete()
            Facture.query.delete()
            Operation.query.delete()
            Client.query.delete()
            Service.query.delete()
            db.session.commit()

        # ── Importer les clients ──
        nb_clients = 0
        for c in data.get('clients', []):
            if mode == 'fusionner' and Client.query.filter_by(numero=c['numero']).first():
                continue
            db.session.add(Client(
                numero=c.get('numero',''), nom=c.get('nom',''),
                adresse=c.get('adresse',''), telephone=c.get('telephone',''),
                email=c.get('email',''), nif=c.get('nif',''), rccm=c.get('rccm',''),
            ))
            nb_clients += 1
        db.session.commit()

        # ── Importer les services ──
        nb_services = 0
        for s in data.get('services', []):
            if mode == 'fusionner' and Service.query.filter_by(libelle=s['libelle'], section=s['section']).first():
                continue
            db.session.add(Service(
                section=s.get('section',''), libelle=s.get('libelle',''),
                prix=s.get('prix', 0), actif=s.get('actif', True),
            ))
            nb_services += 1
        db.session.commit()

        # ── Importer les factures ──
        nb_factures = 0
        for f in data.get('factures', []):
            if mode == 'fusionner' and Facture.query.filter_by(numero=f['numero']).first():
                continue
            date_obj = datetime.strptime(f['date'], '%Y-%m-%d') if f.get('date') else datetime.now()
            db.session.add(Facture(
                numero=f.get('numero',''), date=date_obj,
                nom_client=f.get('nom_client',''), service=f.get('service',''),
                montant_ttc=f.get('montant_ttc', 0), mode_paiement=f.get('mode_paiement',''),
                etat_paiement=f.get('etat_paiement',''), section=f.get('section',''),
                cree_par=f.get('cree_par',''), montant_paye=f.get('montant_paye', 0),
                reste_du=f.get('reste_du', 0),
            ))
            nb_factures += 1
        db.session.commit()

        # ── Importer les lignes de facture ──
        for l in data.get('lignes', []):
            fac = Facture.query.filter_by(numero=l.get('facture_numero')).first()
            if not fac:
                continue
            db.session.add(LigneFacture(
                facture_id=fac.id, service=l.get('service',''),
                prix_unitaire=l.get('prix_unitaire', 0), quantite=l.get('quantite', 1),
                montant_ht=l.get('montant_ht', 0), montant_ttc=l.get('montant_ttc', 0),
            ))
        db.session.commit()

        # ── Importer les opérations ──
        nb_operations = 0
        for o in data.get('operations', []):
            if mode == 'fusionner' and Operation.query.filter_by(numero=o['numero']).first():
                continue
            date_obj = datetime.strptime(o['date'], '%Y-%m-%d') if o.get('date') else datetime.now()
            db.session.add(Operation(
                numero=o.get('numero',''), date=date_obj,
                nom_client=o.get('nom_client',''), service=o.get('service',''),
                montant_ttc=o.get('montant_ttc', 0), type_operation=o.get('type_operation',''),
                categorie=o.get('categorie',''), section=o.get('section',''),
                cree_par=o.get('cree_par',''),
            ))
            nb_operations += 1
        db.session.commit()

        # ── Importer les paiements ──
        nb_paiements = 0
        for p in data.get('paiements', []):
            if mode == 'fusionner' and Paiement.query.filter_by(numero=p['numero']).first():
                continue
            fac = Facture.query.filter_by(numero=p.get('facture_numero')).first()
            if not fac:
                continue
            date_obj = datetime.strptime(p['date'], '%Y-%m-%d') if p.get('date') else datetime.now()
            db.session.add(Paiement(
                numero=p.get('numero',''), date=date_obj,
                facture_id=fac.id, n_facture=p.get('n_facture',''),
                nom_client=p.get('nom_client',''), montant_facture=p.get('montant_facture', 0),
                montant_paye=p.get('montant_paye', 0), reste_du=p.get('reste_du', 0),
                mode_paiement=p.get('mode_paiement',''), etat_facture=p.get('etat_facture',''),
                notes=p.get('notes',''), cree_par=p.get('cree_par',''),
            ))
            nb_paiements += 1
        db.session.commit()

        flash(
            f'✅ Import réussi ! '
            f'{nb_clients} client(s), {nb_services} service(s), '
            f'{nb_factures} facture(s), {nb_operations} opération(s), '
            f'{nb_paiements} paiement(s) importé(s).',
            'success'
        )

    except Exception as e:
        db.session.rollback()
        flash(f'Erreur lors de l\'import : {str(e)}', 'danger')

    return redirect(url_for('sauvegarde'))


# ================================================================
# ROUTES PROFORMA — à ajouter dans app.py
# ================================================================

@app.route('/factures/proforma', methods=['GET', 'POST'])
def nouvelle_proforma():
    if 'username' not in session:
        return redirect(url_for('login'))
    if session['role'] == 'Lecture seule':
        flash('Accès refusé.', 'danger')
        return redirect(url_for('liste_factures'))

    if request.method == 'POST':
        # Générer numéro proforma unique
        numero = generer_numero_facture('PROF')

        nom_client    = request.form['nom_client']
        mode_paiement = request.form.get('mode_paiement', 'Espece')
        section       = request.form['section']
        tva_taux      = float(request.form.get('tva_taux', 18))
        notes         = request.form.get('notes', '')

        services_f     = request.form.getlist('service[]')
        prix_unitaires = request.form.getlist('prix_unitaire[]')
        quantites      = request.form.getlist('quantite[]')

        montant_ttc_total = 0
        lignes_data = []
        for i in range(len(services_f)):
            svc = services_f[i].strip()
            if not svc: continue
            pu  = float(prix_unitaires[i]) if prix_unitaires[i] else 0
            qty = int(quantites[i]) if quantites[i] else 1
            ht  = pu * qty
            tva = round(ht * tva_taux / 100, 2)
            ttc = ht + tva
            montant_ttc_total += ttc
            lignes_data.append({
                'service': svc, 'prix_unitaire': pu,
                'quantite': qty, 'montant_ht': ht, 'montant_ttc': ttc
            })

        # Créer la facture avec type_operation = 'Proforma'
        facture = Facture(
            numero        = numero,
            nom_client    = nom_client,
            service       = lignes_data[0]['service'] if lignes_data else '',
            montant_ttc   = montant_ttc_total,
            mode_paiement = mode_paiement,
            etat_paiement = 'Non Payer',
            type_operation= 'Proforma',   # ← marqueur proforma
            section       = section,
            cree_par      = session['username'],
            montant_paye  = 0,
            reste_du      = montant_ttc_total,
        )
        db.session.add(facture)
        db.session.flush()

        for ld in lignes_data:
            db.session.add(LigneFacture(
                facture_id    = facture.id,
                service       = ld['service'],
                prix_unitaire = ld['prix_unitaire'],
                quantite      = ld['quantite'],
                montant_ht    = ld['montant_ht'],
                montant_ttc   = ld['montant_ttc'],
            ))
        db.session.commit()

        flash(f'Proforma {numero} créée avec succès !', 'success')
        return redirect(url_for('liste_proformas'))

    # GET
    numero_auto = generer_numero_facture('PROF')

    services_list = Service.query.filter_by(actif=True).order_by(Service.section, Service.libelle).all()

    return render_template('nouvelle_proforma.html',
        username     = session['username'],
        role         = session['role'],
        numero_auto  = numero_auto,
        clients      = Client.query.all(),
        today        = datetime.now().strftime('%Y-%m-%d'),
        services_list= services_list,
    )


@app.route('/proformas')
def liste_proformas():
    if 'username' not in session:
        return redirect(url_for('login'))

    proformas = Facture.query.filter_by(type_operation='Proforma')\
                             .order_by(Facture.date.desc()).all()
    total_ttc = sum(p.montant_ttc for p in proformas)

    return render_template('proformas.html',
        username      = session['username'],
        role          = session['role'],
        proformas     = proformas,
        nb_proformas  = len(proformas),
        nb_en_attente = sum(1 for p in proformas if p.etat_paiement == 'Non Payer'),
        nb_converties = sum(1 for p in proformas if p.etat_paiement == 'Converti'),
        total_ttc     = f"{total_ttc:,.0f}".replace(',', ' '),
    )


@app.route('/proformas/apercu/<int:id>')
def apercu_proforma(id):
    if 'username' not in session:
        return redirect(url_for('login'))

    facture = Facture.query.get_or_404(id)
    client  = Client.query.filter_by(nom=facture.nom_client).first()
    params  = Parametres.query.first()
    lignes  = LigneFacture.query.filter_by(facture_id=facture.id).all()

    if not lignes:
        tva_taux   = 18
        montant_ht = round(facture.montant_ttc / 1.18, 2)
        lignes_display = [{'service': facture.service, 'quantite': 1, 'prix_unitaire': montant_ht, 'montant_ht': montant_ht, 'montant_ttc': facture.montant_ttc}]
    else:
        lignes_display = [{'service': l.service, 'quantite': l.quantite, 'prix_unitaire': l.prix_unitaire, 'montant_ht': l.montant_ht, 'montant_ttc': l.montant_ttc} for l in lignes]

    tva_taux  = 18
    total_ht  = sum(l['montant_ht'] for l in lignes_display)
    total_tva = round(facture.montant_ttc - total_ht, 2)

    return render_template('apercu_proforma.html',
        facture     = facture,
        client      = client,
        params      = params,
        lignes      = lignes_display,
        total_ht    = f"{total_ht:,.0f}".replace(',', ' '),
        total_tva   = f"{total_tva:,.0f}".replace(',', ' '),
        montant_ttc = f"{facture.montant_ttc:,.0f}".replace(',', ' '),
        tva_taux    = tva_taux,
        username    = session['username'],
        role        = session['role'],
    )


@app.route('/proformas/convertir/<int:id>', methods=['POST'])
def convertir_proforma(id):
    if 'username' not in session:
        return redirect(url_for('login'))
    if session['role'] == 'Lecture seule':
        flash('Accès refusé.', 'danger')
        return redirect(url_for('liste_proformas'))

    proforma = Facture.query.get_or_404(id)

    if proforma.type_operation != 'Proforma':
        flash('Ce document n\'est pas une proforma.', 'warning')
        return redirect(url_for('liste_proformas'))

    # Générer numéro de facture
    nouveau_numero = generer_numero_facture('FKSP')

    etat = request.form.get('etat_paiement', 'Non Payer')

    # Calcul montant payé / reste dû
    if etat == 'Payer':
        montant_paye = proforma.montant_ttc
        reste_du     = 0
    elif etat == 'Partiel':
        montant_paye = 0
        reste_du     = proforma.montant_ttc
    else:
        montant_paye = 0
        reste_du     = proforma.montant_ttc

    # ── Créer la NOUVELLE facture ──
    nouvelle_facture = Facture(
        numero        = nouveau_numero,
        nom_client    = proforma.nom_client,
        service       = proforma.service,
        montant_ttc   = proforma.montant_ttc,
        mode_paiement = proforma.mode_paiement,
        etat_paiement = etat,
        type_operation= 'Recettes',
        section       = proforma.section,
        cree_par      = session['username'],
        montant_paye  = montant_paye,
        reste_du      = reste_du,
    )
    db.session.add(nouvelle_facture)
    db.session.flush()

    # Copier les lignes vers la nouvelle facture
    lignes = LigneFacture.query.filter_by(facture_id=proforma.id).all()
    for l in lignes:
        db.session.add(LigneFacture(
            facture_id    = nouvelle_facture.id,
            service       = l.service,
            prix_unitaire = l.prix_unitaire,
            quantite      = l.quantite,
            montant_ht    = l.montant_ht,
            montant_ttc   = l.montant_ttc,
        ))

    # ── Marquer la proforma comme convertie (elle reste dans la liste) ──
    proforma.etat_paiement = 'Converti'

    # Créer l'opération associée
    nb_op  = Operation.query.count() + 1
    op_num = f"OPE-{datetime.now().strftime('%d%m%Y')}-{nb_op:03d}"
    db.session.add(Operation(
        numero        = op_num,
        nom_client    = proforma.nom_client,
        service       = proforma.service,
        montant_ttc   = proforma.montant_ttc,
        type_operation= 'Recettes',
        categorie     = 'Facture',
        section       = proforma.section,
        cree_par      = session['username'],
    ))
    db.session.commit()

    flash(f'Proforma {proforma.numero} convertie en facture {nouveau_numero} !', 'success')
    return redirect(url_for('liste_proformas'))


@app.route('/proformas/supprimer/<int:id>', methods=['POST'])
def supprimer_proforma(id):
    if 'username' not in session:
        return redirect(url_for('login'))
    if session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger')
        return redirect(url_for('liste_proformas'))

    proforma = Facture.query.get_or_404(id)
    numero   = proforma.numero
    LigneFacture.query.filter_by(facture_id=id).delete()
    db.session.delete(proforma)
    db.session.commit()
    flash(f'Proforma {numero} supprimée.', 'warning')
    return redirect(url_for('liste_proformas'))



@app.route('/proformas/modifier/<int:id>', methods=['GET', 'POST'])
def modifier_proforma(id):
    if 'username' not in session:
        return redirect(url_for('login'))
    if session['role'] == 'Lecture seule':
        flash('Accès refusé.', 'danger')
        return redirect(url_for('liste_proformas'))

    proforma = Facture.query.get_or_404(id)

    if proforma.etat_paiement == 'Converti':
        flash('Une proforma convertie ne peut plus être modifiée.', 'warning')
        return redirect(url_for('liste_proformas'))

    if request.method == 'POST':
        tva_taux      = float(request.form.get('tva_taux', 18))
        proforma.nom_client    = request.form['nom_client']
        proforma.mode_paiement = request.form.get('mode_paiement', 'Espece')
        proforma.section       = request.form['section']

        services_f     = request.form.getlist('service[]')
        prix_unitaires = request.form.getlist('prix_unitaire[]')
        quantites      = request.form.getlist('quantite[]')

        # Supprimer anciennes lignes
        LigneFacture.query.filter_by(facture_id=proforma.id).delete()

        montant_ttc_total = 0
        premier_service   = ''
        for i in range(len(services_f)):
            svc = services_f[i].strip()
            if not svc: continue
            pu  = float(prix_unitaires[i]) if prix_unitaires[i] else 0
            qty = int(quantites[i]) if quantites[i] else 1
            ht  = pu * qty
            tva = round(ht * tva_taux / 100, 2)
            ttc = ht + tva
            montant_ttc_total += ttc
            if not premier_service:
                premier_service = svc
            db.session.add(LigneFacture(
                facture_id=proforma.id, service=svc,
                prix_unitaire=pu, quantite=qty,
                montant_ht=ht, montant_ttc=ttc,
            ))

        proforma.service     = premier_service
        proforma.montant_ttc = montant_ttc_total
        proforma.reste_du    = montant_ttc_total
        db.session.commit()

        flash(f'Proforma {proforma.numero} modifiée avec succès !', 'success')
        return redirect(url_for('liste_proformas'))

    # GET — charger les lignes existantes
    lignes        = LigneFacture.query.filter_by(facture_id=proforma.id).all()
    services_list = Service.query.filter_by(actif=True).order_by(Service.section, Service.libelle).all()

    return render_template('modifier_proforma.html',
        username      = session['username'],
        role          = session['role'],
        proforma      = proforma,
        lignes        = lignes,
        clients       = Client.query.all(),
        services_list = services_list,
        today         = datetime.now().strftime('%Y-%m-%d'),
    )


# ================================================================
# ROUTE PROFIL & CHANGEMENT MOT DE PASSE — à ajouter dans app.py
# ================================================================

@app.route('/profil')
def profil():
    if 'username' not in session:
        return redirect(url_for('login'))

    utilisateur = Utilisateur.query.filter_by(username=session['username']).first_or_404()

    return render_template('profil.html',
        username    = session['username'],
        role        = session['role'],
        utilisateur = utilisateur,
    )


@app.route('/profil/changer-mdp', methods=['POST'])
def changer_mdp():
    if 'username' not in session:
        return redirect(url_for('login'))

    ancien_mdp   = request.form.get('ancien_mdp', '').strip()
    nouveau_mdp  = request.form.get('nouveau_mdp', '').strip()
    confirmer    = request.form.get('confirmer_mdp', '').strip()

    utilisateur = Utilisateur.query.filter_by(username=session['username']).first_or_404()

    # Vérifier l'ancien mot de passe
    if not check_password_hash(utilisateur.password, ancien_mdp):
        flash('Mot de passe actuel incorrect.', 'danger')
        return redirect(url_for('profil'))

    # Vérifier que le nouveau est différent
    if nouveau_mdp == ancien_mdp:
        flash('Le nouveau mot de passe doit être différent de l\'ancien.', 'warning')
        return redirect(url_for('profil'))

    # Vérifier la longueur
    if len(nouveau_mdp) < 6:
        flash('Le nouveau mot de passe doit contenir au moins 6 caractères.', 'danger')
        return redirect(url_for('profil'))

    # Vérifier la confirmation
    if nouveau_mdp != confirmer:
        flash('Les deux mots de passe ne correspondent pas.', 'danger')
        return redirect(url_for('profil'))

    # Mettre à jour
    utilisateur.password = generate_password_hash(nouveau_mdp)
    db.session.commit()

    flash('✅ Mot de passe changé avec succès !', 'success')
    return redirect(url_for('profil'))



# ================================================================
# ROUTES AGENDA — à ajouter dans app.py
# ================================================================

# 1) Ajouter Evenement dans l'import de database.py :
#    from database import db, ..., Evenement

# 2) Ajouter dans initialiser_base() :
#    db.create_all()  ← déjà présent, crée automatiquement la table

from database import Evenement
from datetime import date as date_type
import calendar

@app.route('/agenda')
def agenda():
    if 'username' not in session:
        return redirect(url_for('login'))

    # Mois/année à afficher
    mois  = int(request.args.get('mois',  datetime.now().month))
    annee = int(request.args.get('annee', datetime.now().year))

    # Navigation
    if mois == 1:
        mois_prec, annee_prec = 12, annee - 1
    else:
        mois_prec, annee_prec = mois - 1, annee

    if mois == 12:
        mois_suiv, annee_suiv = 1, annee + 1
    else:
        mois_suiv, annee_suiv = mois + 1, annee

    # Événements du mois
    from datetime import date as date_cls
    debut_mois = date_cls(annee, mois, 1)
    nb_jours   = calendar.monthrange(annee, mois)[1]
    fin_mois   = date_cls(annee, mois, nb_jours)

    evenements = Evenement.query.filter(
        Evenement.date >= debut_mois,
        Evenement.date <= fin_mois
    ).order_by(Evenement.date, Evenement.heure_debut).all()

    # Organiser par jour
    evts_par_jour = {}
    for ev in evenements:
        jour = ev.date.day
        if jour not in evts_par_jour:
            evts_par_jour[jour] = []
        evts_par_jour[jour].append(ev)

    # Calendrier : semaines du mois
    cal = calendar.monthcalendar(annee, mois)

    # Stats
    nb_confirmes  = sum(1 for e in evenements if e.statut == 'Confirmé')
    nb_tentatives = sum(1 for e in evenements if e.statut == 'Tentative')
    nb_annules    = sum(1 for e in evenements if e.statut == 'Annulé')

    noms_mois = ['', 'Janvier', 'Février', 'Mars', 'Avril', 'Mai', 'Juin',
                 'Juillet', 'Août', 'Septembre', 'Octobre', 'Novembre', 'Décembre']

    return render_template('agenda.html',
        username      = session['username'],
        role          = session['role'],
        mois          = mois,
        annee         = annee,
        nom_mois      = noms_mois[mois],
        cal           = cal,
        evts_par_jour = evts_par_jour,
        evenements    = evenements,
        mois_prec     = mois_prec,
        annee_prec    = annee_prec,
        mois_suiv     = mois_suiv,
        annee_suiv    = annee_suiv,
        today_day     = datetime.now().day,
        today_month   = datetime.now().month,
        today_year    = datetime.now().year,
        clients       = Client.query.order_by(Client.nom).all(),
        services_list = Service.query.filter_by(actif=True).order_by(Service.section, Service.libelle).all(),
        nb_confirmes  = nb_confirmes,
        nb_tentatives = nb_tentatives,
        nb_annules    = nb_annules,
        nb_total      = len(evenements),
        factures_liables = Facture.query.order_by(Facture.date.desc()).all(),
    )


@app.route('/agenda/nouveau', methods=['POST'])
def nouvel_evenement():
    if 'username' not in session:
        return redirect(url_for('login'))
    if session['role'] == 'Lecture seule':
        flash('Accès refusé.', 'danger')
        return redirect(url_for('agenda'))

    titre      = request.form.get('titre', '').strip()
    date_str   = request.form.get('date', '')
    heure_debut= request.form.get('heure_debut', '').strip()
    heure_fin  = request.form.get('heure_fin', '').strip()
    nom_client = request.form.get('nom_client', '').strip()
    service    = request.form.get('service', '').strip()
    section    = request.form.get('section', '').strip()
    lieu       = request.form.get('lieu', '').strip()
    notes      = request.form.get('notes', '').strip()
    statut     = request.form.get('statut', 'Confirmé')
    facture_id = int(request.form.get('facture_id')) if request.form.get('facture_id') else None,

    if not titre or not date_str:
        flash('Le titre et la date sont obligatoires.', 'danger')
        return redirect(url_for('agenda'))

    date_evt = datetime.strptime(date_str, '%Y-%m-%d').date()

    # Détecter les conflits (même jour, statut Confirmé)
    conflits = Evenement.query.filter(
        Evenement.date == date_evt,
        Evenement.statut == 'Confirmé',
    ).all()

    db.session.add(Evenement(
        titre      = titre,
        date       = date_evt,
        heure_debut= heure_debut,
        heure_fin  = heure_fin,
        nom_client = nom_client,
        service    = service,
        section    = section,
        lieu       = lieu,
        notes      = notes,
        statut     = statut,
        cree_par   = session['username'],
        facture_id = int(request.form.get('facture_id')) if request.form.get('facture_id') else None,
    ))
    db.session.commit()

    if conflits and statut == 'Confirmé':
        flash(f'⚠️ Événement ajouté mais attention : {len(conflits)} autre(s) prestation(s) confirmée(s) ce jour-là !', 'warning')
    else:
        flash(f'✅ Événement "{titre}" ajouté avec succès !', 'success')

    return redirect(url_for('agenda', mois=date_evt.month, annee=date_evt.year))


@app.route('/agenda/supprimer/<int:id>', methods=['POST'])
def supprimer_evenement(id):
    if 'username' not in session:
        return redirect(url_for('login'))

    evt = Evenement.query.get_or_404(id)
    mois  = evt.date.month
    annee = evt.date.year
    titre = evt.titre
    db.session.delete(evt)
    db.session.commit()
    flash(f'Événement "{titre}" supprimé.', 'warning')
    return redirect(url_for('agenda', mois=mois, annee=annee))


@app.route('/agenda/modifier/<int:id>', methods=['POST'])
def modifier_evenement(id):
    if 'username' not in session:
        return redirect(url_for('login'))
    if session['role'] == 'Lecture seule':
        flash('Accès refusé.', 'danger')
        return redirect(url_for('agenda'))

    evt = Evenement.query.get_or_404(id)
    evt.titre       = request.form.get('titre', '').strip()
    evt.heure_debut = request.form.get('heure_debut', '').strip()
    evt.heure_fin   = request.form.get('heure_fin', '').strip()
    evt.nom_client  = request.form.get('nom_client', '').strip()
    evt.service     = request.form.get('service', '').strip()
    evt.section     = request.form.get('section', '').strip()
    evt.lieu        = request.form.get('lieu', '').strip()
    evt.notes       = request.form.get('notes', '').strip()
    evt.statut      = request.form.get('statut', 'Confirmé')
    db.session.commit()
    flash(f'Événement "{evt.titre}" modifié !', 'success')
    return redirect(url_for('agenda', mois=evt.date.month, annee=evt.date.year))


@app.route('/agenda/api/evenements')
def api_evenements():
    """API JSON pour récupérer les événements d'un mois"""
    from flask import jsonify
    mois  = int(request.args.get('mois',  datetime.now().month))
    annee = int(request.args.get('annee', datetime.now().year))
    from datetime import date as date_cls
    import calendar as cal_mod
    nb_jours = cal_mod.monthrange(annee, mois)[1]
    debut = date_cls(annee, mois, 1)
    fin   = date_cls(annee, mois, nb_jours)
    evts  = Evenement.query.filter(Evenement.date >= debut, Evenement.date <= fin).all()
    return jsonify([{
        'id'         : e.id,
        'titre'      : e.titre,
        'date'       : e.date.strftime('%Y-%m-%d'),
        'heure_debut': e.heure_debut or '',
        'heure_fin'  : e.heure_fin or '',
        'nom_client' : e.nom_client or '',
        'service'    : e.service or '',
        'section'    : e.section or '',
        'lieu'       : e.lieu or '',
        'statut'     : e.statut or '',
    } for e in evts])


# ================================================================
# ROUTES TECHNICIENS — à ajouter dans app.py
# ================================================================

# Ajouter dans l'import :
# from database import db, ..., Technicien, EvenementTechnicien

from database import Technicien, EvenementTechnicien

@app.route('/techniciens')
def liste_techniciens():
    if 'username' not in session:
        return redirect(url_for('login'))
    if session['role'] not in ['Administrateur', 'Caissier']:
        flash('Accès refusé.', 'danger')
        return redirect(url_for('dashboard'))

    techniciens = Technicien.query.order_by(Technicien.nom).all()

    # Stats par technicien : nombre de prestations
    for t in techniciens:
        t.nb_prestations = EvenementTechnicien.query.filter_by(technicien_id=t.id).count()

    return render_template('techniciens.html',
        username     = session['username'],
        role         = session['role'],
        techniciens  = techniciens,
        nb_total     = len(techniciens),
        nb_dispos    = sum(1 for t in techniciens if t.statut == 'Disponible'),
        nb_inactifs  = sum(1 for t in techniciens if t.statut == 'Inactif'),
    )


@app.route('/techniciens/nouveau', methods=['POST'])
def nouveau_technicien():
    if 'username' not in session:
        return redirect(url_for('login'))
    if session['role'] not in ['Administrateur']:
        flash('Accès refusé.', 'danger')
        return redirect(url_for('liste_techniciens'))

    nom        = request.form.get('nom','').strip().upper()
    telephone  = request.form.get('telephone','').strip()
    email      = request.form.get('email','').strip()
    specialite = request.form.get('specialite','').strip()
    notes      = request.form.get('notes','').strip()

    if not nom:
        flash('Le nom est obligatoire.', 'danger')
        return redirect(url_for('liste_techniciens'))

    db.session.add(Technicien(
        nom=nom, telephone=telephone, email=email,
        specialite=specialite, notes=notes, statut='Disponible',
    ))
    db.session.commit()
    flash(f'Technicien "{nom}" ajouté avec succès !', 'success')
    return redirect(url_for('liste_techniciens'))


@app.route('/techniciens/modifier/<int:id>', methods=['POST'])
def modifier_technicien(id):
    if 'username' not in session:
        return redirect(url_for('login'))
    if session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger')
        return redirect(url_for('liste_techniciens'))

    t = Technicien.query.get_or_404(id)
    t.nom        = request.form.get('nom','').strip().upper()
    t.telephone  = request.form.get('telephone','').strip()
    t.email      = request.form.get('email','').strip()
    t.specialite = request.form.get('specialite','').strip()
    t.statut     = request.form.get('statut', 'Disponible')
    t.notes      = request.form.get('notes','').strip()
    db.session.commit()
    flash(f'Technicien "{t.nom}" modifié !', 'success')
    return redirect(url_for('liste_techniciens'))


@app.route('/techniciens/supprimer/<int:id>', methods=['POST'])
def supprimer_technicien(id):
    if 'username' not in session:
        return redirect(url_for('login'))
    if session['role'] != 'Administrateur':
        flash('Accès refusé.', 'danger')
        return redirect(url_for('liste_techniciens'))

    t = Technicien.query.get_or_404(id)
    nom = t.nom
    # Supprimer les assignations
    EvenementTechnicien.query.filter_by(technicien_id=id).delete()
    db.session.delete(t)
    db.session.commit()
    flash(f'Technicien "{nom}" supprimé.', 'warning')
    return redirect(url_for('liste_techniciens'))


@app.route('/techniciens/toggle/<int:id>', methods=['POST'])
def toggle_technicien(id):
    if 'username' not in session:
        return redirect(url_for('login'))
    t = Technicien.query.get_or_404(id)
    t.statut = 'Inactif' if t.statut == 'Disponible' else 'Disponible'
    db.session.commit()
    flash(f'Statut de "{t.nom}" mis à jour.', 'success')
    return redirect(url_for('liste_techniciens'))


# ── Assigner techniciens à un événement ──
@app.route('/agenda/assigner/<int:evt_id>', methods=['POST'])
def assigner_techniciens(evt_id):
    if 'username' not in session:
        return redirect(url_for('login'))

    evt = Evenement.query.get_or_404(evt_id)

    # Supprimer les anciennes assignations
    EvenementTechnicien.query.filter_by(evenement_id=evt_id).delete()

    technicien_ids = request.form.getlist('techniciens[]')
    roles          = request.form.getlist('roles[]')

    conflits = []
    for i, tid in enumerate(technicien_ids):
        if not tid: continue
        tid = int(tid)

        # Vérifier conflit : technicien déjà assigné ce jour-là
        conflit = db.session.query(EvenementTechnicien).join(
            Evenement, EvenementTechnicien.evenement_id == Evenement.id
        ).filter(
            Evenement.date == evt.date,
            EvenementTechnicien.technicien_id == tid,
            Evenement.id != evt_id,
        ).first()

        if conflit:
            tech = Technicien.query.get(tid)
            conflits.append(tech.nom if tech else str(tid))

        role = roles[i] if i < len(roles) else ''
        db.session.add(EvenementTechnicien(
            evenement_id=evt_id, technicien_id=tid, role=role
        ))

    db.session.commit()

    from flask import jsonify as _jsonify
    if conflits:
        return _jsonify({'ok': True, 'warning': f'Conflit détecté pour : {", ".join(conflits)}'})
    return _jsonify({'ok': True})


# ── API : disponibilité des techniciens pour une date ──
@app.route('/api/techniciens/disponibles')
def api_techniciens_disponibles():
    from flask import jsonify
    date_str = request.args.get('date', '')
    evt_id   = request.args.get('evt_id', 0, type=int)

    techniciens = Technicien.query.filter_by(statut='Disponible').order_by(Technicien.nom).all()
    result = []

    for t in techniciens:
        occupe = False
        if date_str:
            try:
                date_obj = datetime.strptime(date_str, '%Y-%m-%d').date()
                occupe = db.session.query(EvenementTechnicien).join(
                    Evenement, EvenementTechnicien.evenement_id == Evenement.id
                ).filter(
                    Evenement.date == date_obj,
                    EvenementTechnicien.technicien_id == t.id,
                    Evenement.id != evt_id,
                ).first() is not None
            except:
                pass

        # Rôle déjà assigné à cet événement
        assignation = EvenementTechnicien.query.filter_by(
            evenement_id=evt_id, technicien_id=t.id
        ).first()

        result.append({
            'id'          : t.id,
            'nom'         : t.nom,
            'specialite'  : t.specialite or '',
            'telephone'   : t.telephone or '',
            'occupe'      : occupe,
            'role_actuel' : assignation.role if assignation else '',
            'assigne'     : assignation is not None,
        })

    return jsonify(result)


# ================================================================
# ALERTES INTERNES J-1 / J-3 / J-7
# À coller dans app.py avant le if __name__
# ================================================================

@app.route('/api/alertes')
def api_alertes():
    from flask import jsonify
    from datetime import date, timedelta

    if 'username' not in session:
        return jsonify([])

    aujourd_hui = date.today()
    dans_7j     = aujourd_hui + timedelta(days=7)

    # Événements confirmés dans les 7 prochains jours
    evenements = Evenement.query.filter(
        Evenement.date >= aujourd_hui,
        Evenement.date <= dans_7j,
        Evenement.statut == 'Confirmé'
    ).order_by(Evenement.date).all()

    alertes = []
    for ev in evenements:
        delta = (ev.date - aujourd_hui).days
        if delta == 0:
            niveau = 'danger';  label = "Aujourd'hui !"
        elif delta == 1:
            niveau = 'danger';  label = 'Demain (J-1)'
        elif delta <= 3:
            niveau = 'warning'; label = f'Dans {delta} jours (J-{delta})'
        else:
            niveau = 'info';    label = f'Dans {delta} jours (J-{delta})'

        alertes.append({
            'id':     ev.id,
            'titre':  ev.titre,
            'date':   ev.date.strftime('%d/%m/%Y'),
            'heure':  ev.heure_debut or '',
            'client': ev.nom_client or '',
            'lieu':   ev.lieu or '',
            'delta':  delta,
            'label':  label,
            'niveau': niveau,
        })

    return jsonify(alertes)


# ================================================================
# FICHE PRESTATION PDF — à coller dans app.py avant if __name__
# ================================================================

@app.route('/agenda/fiche-pdf/<int:evt_id>')
def fiche_prestation_pdf(evt_id):
    if 'username' not in session:
        return redirect(url_for('login'))

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_LEFT, TA_RIGHT
    from io import BytesIO

    evt = Evenement.query.get_or_404(evt_id)
    params = Parametres.query.first()
    techniciens_assignes = EvenementTechnicien.query.filter_by(evenement_id=evt_id).all()
    facture_liee = Facture.query.get(evt.facture_id) if evt.facture_id else None

    # Couleurs
    KS_DARK   = colors.HexColor('#1a1a2e')
    KS_RED    = colors.HexColor('#e94560')
    KS_LIGHT  = colors.HexColor('#f8f9fa')
    KS_GRAY   = colors.HexColor('#6b7280')
    KS_GREEN  = colors.HexColor('#00b894')
    KS_BORDER = colors.HexColor('#e5e7eb')

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=1.5*cm, leftMargin=1.5*cm,
        topMargin=1.5*cm, bottomMargin=1.5*cm)

    styles = getSampleStyleSheet()
    story  = []

    # ── Styles personnalisés ──────────────────────────────────────
    h1 = ParagraphStyle('h1', fontSize=20, fontName='Helvetica-Bold',
        textColor=KS_DARK, spaceAfter=4)
    h2 = ParagraphStyle('h2', fontSize=11, fontName='Helvetica-Bold',
        textColor=KS_RED, spaceBefore=14, spaceAfter=6)
    label_style = ParagraphStyle('label', fontSize=8, fontName='Helvetica',
        textColor=KS_GRAY)
    value_style = ParagraphStyle('value', fontSize=10, fontName='Helvetica-Bold',
        textColor=KS_DARK)
    small_style = ParagraphStyle('small', fontSize=8, fontName='Helvetica',
        textColor=KS_GRAY, spaceAfter=2)
    center_style = ParagraphStyle('center', fontSize=9, fontName='Helvetica',
        textColor=KS_GRAY, alignment=TA_CENTER)

    nom_entreprise = params.nom_entreprise if params else 'KS Production'
    slogan = params.slogan if params else 'Studio & Sonorisation'
    tel = params.telephone if params else ''
    email = params.email if params else ''
    adresse = params.adresse if params else ''

    # ── EN-TÊTE ───────────────────────────────────────────────────
    header_data = [[
        Paragraph(f'<font color="#e94560"><b>{nom_entreprise}</b></font>', 
            ParagraphStyle('brand', fontSize=22, fontName='Helvetica-Bold', textColor=KS_RED)),
        Paragraph(f'<b>FICHE DE PRESTATION</b>', 
            ParagraphStyle('title', fontSize=14, fontName='Helvetica-Bold',
            textColor=KS_DARK, alignment=TA_RIGHT)),
    ]]
    header_table = Table(header_data, colWidths=[10*cm, 7.5*cm])
    header_table.setStyle(TableStyle([
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
    ]))
    story.append(header_table)

    # Sous-titre entreprise
    info_data = [[
        Paragraph(f'{slogan}<br/><font color="#9ca3af">{tel} | {email} | {adresse}</font>',
            ParagraphStyle('info', fontSize=8, fontName='Helvetica', textColor=KS_GRAY)),
        Paragraph(f'<font color="#9ca3af">Réf. EVT-{evt.id:04d}</font>',
            ParagraphStyle('ref', fontSize=9, fontName='Helvetica',
            textColor=KS_GRAY, alignment=TA_RIGHT)),
    ]]
    info_table = Table(info_data, colWidths=[10*cm, 7.5*cm])
    info_table.setStyle(TableStyle([('VALIGN', (0,0), (-1,-1), 'TOP')]))
    story.append(info_table)

    story.append(HRFlowable(width='100%', thickness=2, color=KS_RED, spaceAfter=12))

    # ── TITRE PRESTATION ─────────────────────────────────────────
    story.append(Paragraph(evt.titre.upper(), 
        ParagraphStyle('prestTitle', fontSize=18, fontName='Helvetica-Bold',
        textColor=KS_DARK, spaceAfter=4)))

    # Badge statut
    statut_color = KS_GREEN if evt.statut == 'Confirmé' else \
                   colors.HexColor('#f59e0b') if evt.statut == 'Tentative' else \
                   colors.HexColor('#ef4444')
    story.append(Paragraph(f'● {evt.statut}',
        ParagraphStyle('statut', fontSize=10, fontName='Helvetica-Bold',
        textColor=statut_color, spaceAfter=16)))

    # ── INFORMATIONS GÉNÉRALES ────────────────────────────────────
    story.append(Paragraph('INFORMATIONS GÉNÉRALES', h2))

    # Ligne date/heure
    heure_str = ''
    if evt.heure_debut:
        heure_str = evt.heure_debut
        if evt.heure_fin:
            heure_str += f' — {evt.heure_fin}'

    info_gen = [
        [
            [Paragraph('DATE', label_style), Paragraph(evt.date.strftime('%d/%m/%Y'), value_style)],
            [Paragraph('HORAIRE', label_style), Paragraph(heure_str or '—', value_style)],
            [Paragraph('SECTION', label_style), Paragraph(evt.section or '—', value_style)],
        ],
        [
            [Paragraph('CLIENT', label_style), Paragraph(evt.nom_client or '—', value_style)],
            [Paragraph('LIEU', label_style), Paragraph(evt.lieu or '—', value_style)],
            [Paragraph('SERVICE', label_style), Paragraph(evt.service or '—', value_style)],
        ],
    ]

    for row in info_gen:
        row_data = []
        for cell in row:
            row_data.append(Table([[cell[0]], [cell[1]]], colWidths=[5.8*cm]))
        t = Table([row_data], colWidths=[5.8*cm, 5.8*cm, 5.8*cm])
        t.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), KS_LIGHT),
            ('ROUNDEDCORNERS', [6]),
            ('BOX', (0,0), (0,0), 0.5, KS_BORDER),
            ('BOX', (1,0), (1,0), 0.5, KS_BORDER),
            ('BOX', (2,0), (2,0), 0.5, KS_BORDER),
            ('LEFTPADDING', (0,0), (-1,-1), 10),
            ('RIGHTPADDING', (0,0), (-1,-1), 10),
            ('TOPPADDING', (0,0), (-1,-1), 8),
            ('BOTTOMPADDING', (0,0), (-1,-1), 8),
            ('VALIGN', (0,0), (-1,-1), 'TOP'),
        ]))
        story.append(t)
        story.append(Spacer(1, 6))

    # ── NOTES ─────────────────────────────────────────────────────
    if evt.notes:
        story.append(Paragraph('NOTES', h2))
        notes_table = Table([[Paragraph(evt.notes, 
            ParagraphStyle('notes', fontSize=9, fontName='Helvetica',
            textColor=KS_DARK, leading=14))]],
            colWidths=[17.5*cm])
        notes_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,-1), colors.HexColor('#fffbeb')),
            ('BOX', (0,0), (-1,-1), 0.5, colors.HexColor('#f59e0b')),
            ('LEFTPADDING', (0,0), (-1,-1), 12),
            ('RIGHTPADDING', (0,0), (-1,-1), 12),
            ('TOPPADDING', (0,0), (-1,-1), 10),
            ('BOTTOMPADDING', (0,0), (-1,-1), 10),
        ]))
        story.append(notes_table)

    # ── ÉQUIPE TECHNIQUE ─────────────────────────────────────────
    story.append(Paragraph('ÉQUIPE TECHNIQUE', h2))

    if techniciens_assignes:
        tech_header = [
            Paragraph('NOM', ParagraphStyle('th', fontSize=8, fontName='Helvetica-Bold',
                textColor=colors.white)),
            Paragraph('SPÉCIALITÉ', ParagraphStyle('th', fontSize=8, fontName='Helvetica-Bold',
                textColor=colors.white)),
            Paragraph('RÔLE SUR LA PRESTATION', ParagraphStyle('th', fontSize=8, fontName='Helvetica-Bold',
                textColor=colors.white)),
            Paragraph('CONTACT', ParagraphStyle('th', fontSize=8, fontName='Helvetica-Bold',
                textColor=colors.white)),
        ]
        tech_rows = [tech_header]
        for i, at in enumerate(techniciens_assignes):
            t = Technicien.query.get(at.technicien_id)
            if not t: continue
            bg = KS_LIGHT if i % 2 == 0 else colors.white
            tech_rows.append([
                Paragraph(t.nom, ParagraphStyle('td', fontSize=9, fontName='Helvetica-Bold', textColor=KS_DARK)),
                Paragraph(t.specialite or '—', ParagraphStyle('td', fontSize=9, fontName='Helvetica', textColor=KS_GRAY)),
                Paragraph(at.role or '—', ParagraphStyle('td', fontSize=9, fontName='Helvetica', textColor=KS_DARK)),
                Paragraph(t.telephone or '—', ParagraphStyle('td', fontSize=9, fontName='Helvetica', textColor=KS_GRAY)),
            ])

        tech_table = Table(tech_rows, colWidths=[5*cm, 3.5*cm, 5.5*cm, 3.5*cm])
        tech_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), KS_DARK),
            ('ROWBACKGROUNDS', (0,1), (-1,-1), [KS_LIGHT, colors.white]),
            ('GRID', (0,0), (-1,-1), 0.3, KS_BORDER),
            ('LEFTPADDING', (0,0), (-1,-1), 8),
            ('RIGHTPADDING', (0,0), (-1,-1), 8),
            ('TOPPADDING', (0,0), (-1,-1), 7),
            ('BOTTOMPADDING', (0,0), (-1,-1), 7),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        story.append(tech_table)
    else:
        story.append(Paragraph('Aucun technicien assigné à cette prestation.',
            ParagraphStyle('empty', fontSize=9, fontName='Helvetica', textColor=KS_GRAY,
            spaceAfter=8)))

    # ── FACTURE LIÉE ─────────────────────────────────────────────
    if facture_liee:
        story.append(Paragraph('FACTURE LIÉE', h2))
        etat_color = KS_GREEN if facture_liee.etat_paiement == 'Payer' else \
                     colors.HexColor('#f59e0b') if facture_liee.etat_paiement == 'Partiel' else \
                     colors.HexColor('#ef4444')
        etat_txt = 'PAYÉE' if facture_liee.etat_paiement == 'Payer' else \
                   'PARTIELLE' if facture_liee.etat_paiement == 'Partiel' else 'IMPAYÉE'

        fac_data = [
            [Paragraph('N° FACTURE', label_style), Paragraph('CLIENT', label_style),
             Paragraph('MONTANT TTC', label_style), Paragraph('ÉTAT', label_style)],
            [
                Paragraph(facture_liee.numero, value_style),
                Paragraph(facture_liee.nom_client, value_style),
                Paragraph(f"{facture_liee.montant_ttc:,.0f} FCFA".replace(',', ' '), value_style),
                Paragraph(etat_txt, ParagraphStyle('etat', fontSize=10, fontName='Helvetica-Bold',
                    textColor=etat_color)),
            ]
        ]
        fac_table = Table(fac_data, colWidths=[4.5*cm, 5*cm, 4.5*cm, 3.5*cm])
        fac_table.setStyle(TableStyle([
            ('BACKGROUND', (0,0), (-1,0), KS_LIGHT),
            ('BACKGROUND', (0,1), (-1,1), colors.white),
            ('BOX', (0,0), (-1,-1), 0.5, KS_BORDER),
            ('INNERGRID', (0,0), (-1,-1), 0.3, KS_BORDER),
            ('LEFTPADDING', (0,0), (-1,-1), 10),
            ('RIGHTPADDING', (0,0), (-1,-1), 10),
            ('TOPPADDING', (0,0), (-1,-1), 7),
            ('BOTTOMPADDING', (0,0), (-1,-1), 7),
            ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ]))
        story.append(fac_table)

    # ── PIED DE PAGE ─────────────────────────────────────────────
    story.append(Spacer(1, 20))
    story.append(HRFlowable(width='100%', thickness=0.5, color=KS_BORDER, spaceAfter=8))
    from datetime import datetime as dt
    story.append(Paragraph(
        f'Document généré le {dt.now().strftime("%d/%m/%Y à %H:%M")} — {nom_entreprise} | {tel} | {email}',
        ParagraphStyle('footer', fontSize=7, fontName='Helvetica',
        textColor=KS_GRAY, alignment=TA_CENTER)))

    # ── Génération ────────────────────────────────────────────────
    doc.build(story)
    buffer.seek(0)
    from flask import send_file
    titre_safe = evt.titre.replace(' ', '_').replace('/', '-')[:30]
    filename = f"Prestation_{titre_safe}_{evt.date.strftime('%d%m%Y')}.pdf"
    import base64
    pdf_b64 = base64.b64encode(buffer.read()).decode('utf-8')
    return render_template('apercu_prestation_pdf.html',
        pdf_b64=pdf_b64,
        filename=filename,
        evt=evt)

# ================================================================
# RELANCES IMPAYÉS — à coller dans app.py avant if __name__
# ================================================================

@app.route('/relances')
def relances_impayes():
    if 'username' not in session:
        return redirect(url_for('login'))

    from datetime import date, timedelta
    from flask import jsonify

    seuil_jours = int(request.args.get('jours', 30))
    date_limite = date.today() - timedelta(days=seuil_jours)

    # Factures impayées ou partielles depuis + de X jours
    factures_relance = Facture.query.filter(
        Facture.etat_paiement.in_(['Non Payer', 'Partiel']),
        Facture.date <= datetime.combine(date_limite, datetime.min.time())
    ).order_by(Facture.date.asc()).all()

    # Calculer les jours de retard pour chaque facture
    relances = []
    for f in factures_relance:
        jours_retard = (date.today() - f.date.date()).days
        relances.append({
            'facture':      f,
            'jours_retard': jours_retard,
            'urgence':      'haute' if jours_retard > 60 else 'moyenne' if jours_retard > 30 else 'basse'
        })

    params = Parametres.query.first()

    return render_template('relances.html',
        username    = session['username'],
        role        = session['role'],
        relances    = relances,
        seuil_jours = seuil_jours,
        nb_total    = len(relances),
        total_impaye= sum(r['facture'].reste_du for r in relances),
        params      = params,
    )


@app.route('/relances/pdf/<int:facture_id>')
def relance_pdf(facture_id):
    if 'username' not in session:
        return redirect(url_for('login'))

    from reportlab.lib.pagesizes import A4
    from reportlab.lib import colors
    from reportlab.lib.units import cm
    from reportlab.platypus import SimpleDocTemplate, Paragraph, Spacer, Table, TableStyle, HRFlowable
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    from reportlab.lib.enums import TA_CENTER, TA_RIGHT, TA_LEFT
    from io import BytesIO
    from datetime import date

    facture = Facture.query.get_or_404(facture_id)
    params  = Parametres.query.first()

    jours_retard = (date.today() - facture.date.date()).days

    KS_DARK   = colors.HexColor('#1a1a2e')
    KS_RED    = colors.HexColor('#e94560')
    KS_LIGHT  = colors.HexColor('#f8f9fa')
    KS_GRAY   = colors.HexColor('#6b7280')
    KS_ORANGE = colors.HexColor('#f59e0b')
    KS_BORDER = colors.HexColor('#e5e7eb')

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=A4,
        rightMargin=2*cm, leftMargin=2*cm,
        topMargin=2*cm, bottomMargin=2*cm)

    styles = getSampleStyleSheet()
    story  = []

    nom_entreprise = params.nom_entreprise if params else 'KS Production'
    tel            = params.telephone if params else ''
    email          = params.email if params else ''
    adresse        = params.adresse if params else ''

    h2 = ParagraphStyle('h2', fontSize=11, fontName='Helvetica-Bold',
        textColor=KS_RED, spaceBefore=16, spaceAfter=8)
    body = ParagraphStyle('body', fontSize=10, fontName='Helvetica',
        textColor=KS_DARK, leading=16)
    label_s = ParagraphStyle('label', fontSize=8, fontName='Helvetica', textColor=KS_GRAY)
    value_s = ParagraphStyle('value', fontSize=10, fontName='Helvetica-Bold', textColor=KS_DARK)

    # ── EN-TÊTE ───────────────────────────────────────────────────
    header = Table([[
        Paragraph(f'<font color="#e94560"><b>{nom_entreprise}</b></font>',
            ParagraphStyle('brand', fontSize=20, fontName='Helvetica-Bold', textColor=KS_RED)),
        Paragraph('<b>LETTRE DE RELANCE</b>',
            ParagraphStyle('title', fontSize=13, fontName='Helvetica-Bold',
            textColor=KS_DARK, alignment=TA_RIGHT)),
    ]], colWidths=[9*cm, 8*cm])
    header.setStyle(TableStyle([('VALIGN',(0,0),(-1,-1),'MIDDLE')]))
    story.append(header)

    story.append(Paragraph(
        f'{tel} | {email} | {adresse}',
        ParagraphStyle('info', fontSize=8, fontName='Helvetica', textColor=KS_GRAY, spaceAfter=4)))
    story.append(HRFlowable(width='100%', thickness=2, color=KS_RED, spaceAfter=16))

    # Date et référence
    today_str = date.today().strftime('%d/%m/%Y')
    story.append(Table([[
        Paragraph(f'Lomé, le {today_str}',
            ParagraphStyle('date', fontSize=10, fontName='Helvetica', textColor=KS_GRAY)),
        Paragraph(f'Réf. RELANCE-{facture.numero}',
            ParagraphStyle('ref', fontSize=10, fontName='Helvetica-Bold',
            textColor=KS_DARK, alignment=TA_RIGHT)),
    ]], colWidths=[8.5*cm, 8.5*cm]))
    story.append(Spacer(1, 20))

    # Destinataire
    story.append(Paragraph(f'<b>À l\'attention de :</b>', body))
    story.append(Paragraph(f'<b>{facture.nom_client}</b>',
        ParagraphStyle('client', fontSize=12, fontName='Helvetica-Bold',
        textColor=KS_DARK, spaceAfter=20)))

    # Objet
    story.append(Table([[
        Paragraph(f'Objet : Relance pour facture impayée — {facture.numero}',
            ParagraphStyle('objet', fontSize=10, fontName='Helvetica-Bold', textColor=KS_DARK)),
    ]], colWidths=[17*cm]))
    story.append(Spacer(1, 16))

    # Corps de la lettre
    intro = f"""Madame, Monsieur,

Sauf erreur de notre part, nous constatons que la facture n° <b>{facture.numero}</b> 
d'un montant de <b>{facture.montant_ttc:,.0f} FCFA</b>, émise le <b>{facture.date.strftime('%d/%m/%Y')}</b>, 
reste à ce jour <b>impayée depuis {jours_retard} jour(s)</b>."""

    story.append(Paragraph(intro.replace('\n', '<br/>'),
        ParagraphStyle('body', fontSize=10, fontName='Helvetica', textColor=KS_DARK, leading=18, spaceAfter=12)))

    # Détail facture
    story.append(Paragraph('DÉTAIL DE LA CRÉANCE', h2))
    fac_rows = [
        [Paragraph('N° FACTURE', label_s), Paragraph('DATE', label_s),
         Paragraph('MONTANT TTC', label_s), Paragraph('DÉJÀ PAYÉ', label_s),
         Paragraph('RESTE DÛ', label_s)],
        [
            Paragraph(facture.numero, value_s),
            Paragraph(facture.date.strftime('%d/%m/%Y'), value_s),
            Paragraph(f"{facture.montant_ttc:,.0f} FCFA".replace(',', ' '), value_s),
            Paragraph(f"{facture.montant_paye:,.0f} FCFA".replace(',', ' '),
                ParagraphStyle('v2', fontSize=10, fontName='Helvetica-Bold',
                textColor=colors.HexColor('#00b894'))),
            Paragraph(f"{facture.reste_du:,.0f} FCFA".replace(',', ' '),
                ParagraphStyle('v3', fontSize=11, fontName='Helvetica-Bold',
                textColor=KS_RED)),
        ]
    ]
    fac_table = Table(fac_rows, colWidths=[3.5*cm, 2.8*cm, 3.5*cm, 3.2*cm, 3*cm] if facture.montant_paye > 0 else [4*cm, 3.5*cm, 4*cm, 2.5*cm, 3*cm])
    fac_table.setStyle(TableStyle([
        ('BACKGROUND', (0,0), (-1,0), KS_LIGHT),
        ('BOX', (0,0), (-1,-1), 0.5, KS_BORDER),
        ('INNERGRID', (0,0), (-1,-1), 0.3, KS_BORDER),
        ('LEFTPADDING', (0,0), (-1,-1), 10),
        ('RIGHTPADDING', (0,0), (-1,-1), 10),
        ('TOPPADDING', (0,0), (-1,-1), 8),
        ('BOTTOMPADDING', (0,0), (-1,-1), 8),
        ('VALIGN', (0,0), (-1,-1), 'MIDDLE'),
        ('BACKGROUND', (-1,1), (-1,1), colors.HexColor('#fff5f7')),
    ]))
    story.append(fac_table)
    story.append(Spacer(1, 16))

    # Message de relance
    message = f"""Nous vous remercions de bien vouloir procéder au règlement de la somme de 
<b>{facture.reste_du:,.0f} FCFA</b> dans les meilleurs délais.

Pour tout règlement ou toute question concernant cette facture, n'hésitez pas à nous 
contacter aux coordonnées mentionnées ci-dessus.

Dans le cas où ce règlement aurait déjà été effectué, veuillez considérer cette lettre 
comme sans effet et nous en excuser."""

    story.append(Paragraph(message.replace('\n', '<br/>'),
        ParagraphStyle('body2', fontSize=10, fontName='Helvetica', textColor=KS_DARK,
        leading=18, spaceAfter=24)))

    # Formule de politesse
    story.append(Paragraph(
        'Dans l\'attente de votre règlement, nous vous adressons nos cordiales salutations.',
        ParagraphStyle('body3', fontSize=10, fontName='Helvetica', textColor=KS_DARK,
        leading=16, spaceAfter=30)))

    # Signature
    story.append(Table([[
        Paragraph('', body),
        Table([[
            Paragraph(f'<b>{nom_entreprise}</b>',
                ParagraphStyle('sig', fontSize=11, fontName='Helvetica-Bold', textColor=KS_DARK)),
            Paragraph(f'{tel}<br/>{email}',
                ParagraphStyle('sig2', fontSize=9, fontName='Helvetica', textColor=KS_GRAY)),
        ]], colWidths=[8*cm]),
    ]], colWidths=[9*cm, 8*cm]))

    # Pied de page
    story.append(Spacer(1, 30))
    story.append(HRFlowable(width='100%', thickness=0.5, color=KS_BORDER, spaceAfter=6))
    story.append(Paragraph(
        f'Document généré le {today_str} — {nom_entreprise} | {tel} | {email}',
        ParagraphStyle('footer', fontSize=7, fontName='Helvetica',
        textColor=KS_GRAY, alignment=TA_CENTER)))

    doc.build(story)
    buffer.seek(0)

    import base64
    pdf_b64  = base64.b64encode(buffer.read()).decode('utf-8')
    filename = f"Relance_{facture.numero}_{date.today().strftime('%d%m%Y')}.pdf"

    return render_template('apercu_relance_pdf.html',
        pdf_b64=pdf_b64, filename=filename, facture=facture)


@app.route('/relances/message/<int:facture_id>')
def relance_message(facture_id):
    from flask import jsonify
    from datetime import date

    facture = Facture.query.get_or_404(facture_id)
    params  = Parametres.query.first()
    nom_ent = params.nom_entreprise if params else 'KS Production'
    tel     = params.telephone if params else ''

    jours = (date.today() - facture.date.date()).days
    today_str = date.today().strftime('%d/%m/%Y')

    message = f"""Bonjour {facture.nom_client},

Nous vous contactons au sujet de la facture n° {facture.numero} d'un montant de {facture.montant_ttc:,.0f} FCFA, émise le {facture.date.strftime('%d/%m/%Y')}.

À ce jour, un montant de {facture.reste_du:,.0f} FCFA reste impayé depuis {jours} jour(s).

Nous vous prions de bien vouloir régulariser cette situation dans les meilleurs délais.

Pour tout renseignement, contactez-nous au {tel}.

Cordialement,
{nom_ent}"""

    return jsonify({'ok': True, 'message': message.replace(',', ' ').replace(',,', ',')})


with app.app_context():
    initialiser_base()

if __name__ == '__main__':
    app.run(debug=False)
